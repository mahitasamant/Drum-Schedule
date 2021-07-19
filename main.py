
import tkinter #for logo image

import os 
from tkinter import * #for GUI
from tkinter import ttk #for GUI
from PIL import Image, ImageTk #for logo image
import pandas as pd
import math as m
from tkinter import filedialog

from tkinter import messagebox
from tkinter.ttk import Combobox
from openpyxl import load_workbook 
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment  
import os
from tkinter.filedialog import askdirectory

maxcablesize=300
maxmotor=2000
cb7a=''
cb8a=''
cb9a=''
cb10a=''
cb11a=''
cb12a=''
system=''
#cb1=''
f=''
nrows=0
f1=''
f2=''
nrows1=0
cablesizing=''
drumschedule=''
cableschedule=''
Dsa1=1
Dma1=1
Dsg=1
Dmg=1
INPUT1=1
INPUT2=1
INPUT3=1
INPUT4=1
INPUT5=1
INPUT6=1
INPUT7=1
INPUT8=1
INPUT9=1
INPUT10=1
INPUT11=1
INPUT12=1
calc1=''
calc2=''
calc3=''
calc4=''
calc5=''
calc6=''
calc7=''
calc8=''
form5=''
form6=''
form7=''
form8='' 
reset=''
reset1=''
browseButton_Excel=''
browseButton_Excel1=''
calccablesize1=''
calccablesize=''
df=''
df1=''
go14=''
#var=1
#var1=1
#vendor='Polycab'
cbalum=''
factor1=5
factor2=1.5
factor3=1
factor4=3



mainwindow=Tk()
mainwindow.geometry('1380x680+5+5')
mainwindow.title('CABLE MANAGER HOME')
#mainwindow.attributes('-fullscreen', True)
mainwindow.configure(bg='gainsboro')
heading=Label(mainwindow,bg='gainsboro', fg='black',font=('arial', 30,'bold'),text='CABLE MANAGER')
heading.pack()
heading1=Label(mainwindow,bg='gainsboro', fg='black',font=('arial', 30,'bold'),text='Select Your Option:')
heading1.place(x=100,y=100)

dbaddr=""
faddr=""

logo = Image.open("Test.jpg") #L&T Logo
logo1 = logo.resize((250, 250), Image.ANTIALIAS)
logo2 = ImageTk.PhotoImage(logo1)
logo3 = tkinter.Label(image=logo2)
logo3.image = logo2
#logo3.pack(side=TOP)
try:
    def design(x):    
        global faddr
        global vendor
        global f2
        wb=openpyxl.load_workbook(f2+'/output/{}.xlsx'.format(x))
        ws=wb['Sheet1']
        redFill = PatternFill(start_color='008080',
                           end_color='008080',
                           fill_type='solid')
        
        blueFill = PatternFill(start_color='ffcccb',
                           end_color='ffcccb',
                           fill_type='solid')
        #a=ws['F2'].value
        #print(a)
        #print(ws.max_row)
        for i in range(2,ws.max_row+1):
            
            if ws['AK{}'.format(i)].value!='-' and ws['AK{}'.format(i)].value!='Incomplete Data Provided':
                #print('true')
                for cell in ws['A{0}:AK{0}'.format(i)]:
                    for k in range(37):
                        cell[k].fill = redFill
        
            if ws['AK{}'.format(i)].value=='Incomplete Data Provided':
                #print('true')
                for cell in ws['A{0}:AK{0}'.format(i)]:
                    for k in range(37):
                        cell[k].fill = blueFill
        
        
        
        
        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))
        
        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width
            
        for i in range(2,ws.max_row+1):
            for j in range(1,ws.max_column+1):
                cell = ws.cell(row=i, column=j)  
                cell.alignment = Alignment(horizontal='center', vertical='center')    
                
        wb.save(f2+'/output/HV_Cable Sizing_{}.xlsx'.format(vendor))

    def myfunc(df):    
        index=df.index
        rows=len(index)
        #print(rows)
        res={}
        for i in range(1,rows):
            try:
            #print(df.at[i,'Power Cable Size'])
                new=df.at[i,'Power Cable Size'] +' ' + df.at[i,'Type of Conductor']
                if new in res:
                    if int(df.at[i,'Power Cable Size'][0])==3:
                        res.update({new:res[new]+df.at[i,'Cable Length']*df.at[i,'No of runs']})
                    else:
                        res.update({new:res[new]+df.at[i,'Cable Length']*3*df.at[i,'No of runs']})
                else:
                    if int(df.at[i,'Power Cable Size'][0])==3:
                        res.update({new:df.at[i,'Cable Length']*df.at[i,'No of runs']})
                    else:
                        res.update({new:df.at[i,'Cable Length']*3*df.at[i,'No of runs']})
            except:
                continue
        #print(res)
        #res1=list(res)
        #print(res1)
        a=list(res.keys())
        b=list(res.values())
        #df1=pd.DataFrame(data=res,index=None,columns=['Cable  Size','Cable Length'])
        data1={'Cable Size':a,'Cable Length':b}
        df1=pd.DataFrame(data=data1,index=None)
        return(df1)
    
    def cableGlands(df6):    
        index=df6.index
        rows=len(index)
        #print(rows)
        res={}
        for i in range(1,rows):
            try:
                new=df6.at[i,'Gland Size']
                if new in res:
                    res.update({new:res[new]+int(df6.loc[i,'Glands Quantity'])})
                else:
                    res.update({new:int(df6.loc[i,'Glands Quantity'])})       
            except:
                continue
        a=list(res.keys())
        b=list(res.values())
        #df1=pd.DataFrame(data=res,index=None,columns=['Cable  Size','Cable Length'])
        data1={'Cable Gland Size':a,'Gland Quantity':b}
        df7=pd.DataFrame(data=data1,index=None)
        print(df7)
        return(df7)
    

    def formatoutmotor(fault,time,conductor,Tag,rating,voltage,ampacity,derating,cable,runningVD,startingVD,resistance,reactance,runs,length,size,runpf,startpf,efficiency):
        global dbaddr
        global f2
        global system
        wb = load_workbook(dbaddr+"File format\\Motor.xlsx")  
        sheet = wb.active 
        '''
        fault=25
        time=0.3
        k=0.094
        Tag='KM-3702 A'
        rating=315
        voltage=3300
        ampacity=295
        derating=0.63
        cable='3C x 150'
        runningVD=0.04
        startingVD=0.10
        resistance=0.264
        reactance=0.076
        runs=1
        length=260
        size=cable[5:]
        efficiency=0.96
        runpf=0.8
        startpf=0.3
        '''
        if conductor=='AL':
            k=0.094
            sheet['G18'] ='(For Aluminium conductor)'
        else:
            k=0.143
            sheet['G18'] ='(For Copper conductor)'
        sheet['K5'] = '=Today()'  
        sheet['E15'] = fault
        sheet['E16'] = time
        sheet['G16'] = '( Tripping time is considered as {}s)'.format(time)
        sheet['E18'] = k
        sheet['C21'] = 'Nearest higher size of cable is {} mm²'.format(size)
        sheet['F26'] = Tag
        sheet['F27'] = rating
        sheet['F28'] = voltage
        sheet['F29'] = runpf
        sheet['F30'] = efficiency
        sheet['F31'] = startpf
        sheet['C37'] = 'Ampacity of {} Sq.mm Al Cable'.format(size)
        sheet['F37'] = ampacity
        sheet['H40'] = derating
        reqruns=m.ceil(rating*1000/(1.732*voltage*runpf*efficiency*derating*ampacity))
        sheet['H42'] = str(reqruns) +'R x ' + str(cable)
        sheet['F45'] = runningVD
        sheet['F46'] = startingVD
        sheet['C47'] = 'Data for {} Sq.mm cable are as under'.format(cable)
        sheet['F49'] = resistance
        sheet['F50'] = reactance
        sheet['F51'] = runs
        sheet['F52'] = length
        sheet['C71'] = 'Based on short circuit withstand capacity, the minimum size required is {} sq.mm. After checking'.format(size)
        #sheet['C72'] = 'the voltage drop & ampacity, {0} run of {1} sq.mm {2} cable is required.'.format(runs,cable,conductor)
        #sheet['C74'] = 'Hence the selected cable is of voltage grade {0} volts ({3}). {1} Run of {2} Sq.mm. {4} conductor,'.format(voltage,runs,cable,system,conductor)
        if int(cable[0])==3:
            sheet['C72'] = 'the voltage drop & ampacity, {0} run of {1} sq.mm {2} cable is required.'.format(runs,cable,conductor)
            sheet['C74'] = 'Hence the selected cable is of voltage grade {0} volts ({3}). {1} Run of {2} Sq.mm. {4} conductor,'.format(voltage,runs,cable,system,conductor)
        else:
            sheet['C72'] = 'the voltage drop & ampacity, Per Phase {0} run of {1} sq.mm {2} cable is required.'.format(runs,cable,conductor)
            sheet['C74'] = 'Hence the selected cable is of voltage grade {0} volts ({3}). Per Phase {1} Run of {2} Sq.mm. {4} conductor,'.format(voltage,runs,cable,system,conductor)
        
        
          
          
        #now = time.strftime("%x")  
        #sheet['A5'] = now  
          
        wb.save(f2+'/output/{}.xlsx'.format(Tag))  
        
    def formatoutfeeder(fault,time,conductor,Tag,rating,voltage,ampacity,derating,cable,runningVD,resistance,reactance,runs,length,size,runpf):
        global dbaddr
        global f2
        global system
        wb = load_workbook(dbaddr+"File format\\Power.xlsx")  
        sheet = wb.active 
        
        if conductor=='AL':
            k=0.094
            sheet['G17'] ='(For Aluminium conductor)'
        else:
            k=0.143
            sheet['G17'] ='(For Copper conductor)'
        sheet['K5'] = '=Today()'  
        sheet['E15'] = fault
        sheet['E16'] = time
        sheet['G16'] = '( Tripping time is considered as {}s)'.format(time)
        sheet['E17'] = k
        sheet['C20'] = 'Nearest higher size of cable is {} mm²'.format(size)
        sheet['F25'] = Tag
        sheet['F26'] = rating
        sheet['F27'] = voltage
        sheet['F28'] = runpf
        #sheet['F30'] = efficiency
        #sheet['F31'] = startpf
        sheet['C32'] = 'Ampacity of {} Sq.mm Al Cable'.format(size)
        sheet['F32'] = ampacity
        sheet['H35'] = derating
        reqruns=m.ceil(rating*1000/(1.732*voltage*derating*ampacity))
        sheet['H37'] = str(reqruns) +'R x ' + str(cable)
        sheet['F40'] = runningVD
        #sheet['F46'] = startingVD
        sheet['C42'] = 'Data for {} Sq.mm cable are as under'.format(cable)
        sheet['F44'] = resistance
        sheet['F45'] = reactance
        sheet['F46'] = runs
        sheet['F47'] = length
        sheet['C58'] = 'Based on short circuit withstand capacity, the minimum size required is {} sq.mm. After checking'.format(size)
        #print(cable)
        if int(cable[0])==3:
            sheet['C59'] = 'the voltage drop & ampacity, {0} run of {1} sq.mm {2} cable is required.'.format(runs,cable,conductor)
            sheet['C61'] = 'Hence the selected cable is of voltage grade {0} volts ({3}). {1} Run of {2} Sq.mm. {4} conductor,'.format(voltage,runs,cable,system,conductor)
        else:
            sheet['C59'] = 'the voltage drop & ampacity, Per Phase {0} run of {1} sq.mm {2} cable is required.'.format(runs,cable,conductor)
            sheet['C61'] = 'Hence the selected cable is of voltage grade {0} volts ({3}). Per Phase {1} Run of {2} Sq.mm. {4} conductor,'.format(voltage,runs,cable,system,conductor)
        
               
          
        #now = time.strftime("%x")  
        #sheet['A5'] = now  
          
        wb.save(f2+'/Output/{}.xlsx'.format(Tag))  
    
    def HVmotor(faultcurrent,time,rating,voltage,runningpowerfactor,startingpowerfactor,motorefficiency,dfactor,length,mstart):
        ty={}
        acceptable=[]
        acceptable1=[]
        acceptable2=[]
        acceptable3=[]
        acceptable4=[]
        finalset=[]
        #conductor='AL'
        global df1
        global i
        global nrows1
        global system
        global vendor
        global cbalum
        global factor1
        global factor2
        global factor3
        global factor4
        #global cb1
        #importing cable data in dataframe
        df=pd.DataFrame()
        f=dbaddr+vendor+'\\'+'Cable data.xlsx'
        data=pd.read_excel(f,str(voltage)+system)
        df=df.append(data)
        index=df.index
        nrows=len(index)
        #print(cb1)
        fullloadcurrent=float("{:.2f}". format(rating*1000/(1.732*voltage*runningpowerfactor*motorefficiency)))
        if mstart.upper()=='DOL':
            factor=factor1
            #print('DOL ',factor1)
        elif mstart.upper()=='VFD':
            factor=factor2
            #print('VFD ',factor2)
        elif mstart.upper()=='Softstarter'.upper():
            factor=factor3
            #print('SS ',factor3)
        elif mstart.upper()=='StarDelta'.upper():
            factor=factor4
            #print('SD ',factor4)
        
        startingcurrent=float("{:.2f}". format(fullloadcurrent*factor*1.2))        
        
        #checkl- conductor size based on short circuit level
        
        
        sinPF=float("{:.2f}". format(m.sqrt(1-runningpowerfactor**2)))
        sinPFstart=float("{:.2f}". format(m.sqrt(1-startingpowerfactor**2)))
        #conductor='Al'
        
        kal=0.094
        kcu=0.143
            
        areaal=float("{:.2f}". format(faultcurrent*m.sqrt(time)/kal))
        areacu=float("{:.2f}". format(faultcurrent*m.sqrt(time)/kcu))

        
        
        
        if maxmotor<=rating:
            deratingfactor=dfactor[0]
            for k in range(1,nrows):
                if df.at[k,'Conductor']=='AL':
                    
                    if df.at[k,'Core']==1 and df.at[k,'Size']>=areaal:
                        acceptable.append(k)
                else:
                    if cbalum!='YES':
                        if df.at[k,'Core']==1 and df.at[k,'Size']>=areacu:
                            acceptable.append(k)
        
        else:
            #deratingfactor=dfactor[1]
            for k in range(1,nrows):
                #print(df.at[k,'Size'])
                if df.at[k,'Conductor']=='AL':
                    
                    if df.at[k,'Core']==3 and df.at[k,'Size']>=areaal and df.at[k,'Size']<=maxcablesize:
                        acceptable.append(k)
                        deratingfactor=dfactor[1]
                    elif df.at[k,'Core']==1 and df.at[k,'Size']>=areaal and df.at[k,'Size']>maxcablesize:
                        acceptable.append(k)
                        deratingfactor=dfactor[0]
                    
                else:
                    if cbalum!='YES':
                        if df.at[k,'Core']==3 and df.at[k,'Size']>=areacu and df.at[k,'Size']<=maxcablesize:
                            acceptable.append(k)
                            deratingfactor=dfactor[1]
                        elif df.at[k,'Core']==1 and df.at[k,'Size']>=areacu and df.at[k,'Size']>maxcablesize:
                            acceptable.append(k)
                            deratingfactor=dfactor[0]
        #print(acceptable)
        for noofruns in range(1,6):
        #check2- filter cables based on ampacity    
            for l in acceptable:
                ampacityofselectedcable=df.at[l,dfactor[2]]
                
                #deratingfactor=dfactor[1]
                designcurrent=float("{:.2f}". format(noofruns*ampacityofselectedcable*deratingfactor))        
                if designcurrent>=fullloadcurrent:
                    acceptable1.append(l)
            #print('acceptable1',acceptable1)      
            PermissibleRunningVoltageDrop = df1.at[i,'Permissible Running VD']
            
        #check3- filter cables based on running voltage drop    
            for p in acceptable1:
                #noofruns=m.ceil(fullloadcurrent/ty[p][0])
                resistance=df.at[p,'R']
                reactance=df.at[p,'X']
                runningvoltagedrop=1.732*(fullloadcurrent/noofruns)*(length/1000)*(resistance*runningpowerfactor+reactance*sinPF)
                #print(runningvoltagedrop/33)
                if runningvoltagedrop*100/voltage<=PermissibleRunningVoltageDrop:
                    acceptable2.append(p)
            #print('acceptable2',acceptable2)    
            PermissibleStartingVoltageDrop = df1.at[i,'Permissible Starting VD']
            
        #check4- filter cables based on starting voltage drop        
            for q in acceptable1:
                #noofruns=m.ceil(fullloadcurrent/ty[q][0])
                resistance=df.at[q,'R']
                reactance=df.at[q,'X']
                startingvoltagedrop=1.732*(startingcurrent/noofruns)*(length/1000)*(resistance*startingpowerfactor+reactance*sinPFstart)
                #print(runningvoltagedrop/110)
                if startingvoltagedrop*100/voltage<=PermissibleStartingVoltageDrop:
                    acceptable3.append(q)
            #print('acceptable3',acceptable3)   
            
            
        #Create a tuple with all possible cable sizes for that rating
            try:
                if acceptable3!=[]:
                    finalset.append((acceptable3,noofruns))
            except:
                pass
            #finalset=list(set(finalset))
            #print('finalset',finalset)
            acceptable1=[]
            acceptable2=[]
            acceptable3=[]
            acceptable4=[]
        
        #print(1, finalset)
        
        #initialize a selected cable size based on conductor material used
        reqarr={'req':(df.at[finalset[0][0][0],'Size']*finalset[0][1],finalset[0][0][0],finalset[0][1])}
        #print(reqarr)
        
        
        #check4- select cable size with minimum conductor size
        for rank in finalset:
            #print(rank)
            for inde in rank[0]:
                if df.at[inde,'Size']*rank[1]<reqarr['req'][0]:
                    reqarr.update({'req':(df.at[inde,'Size']*rank[1],inde,rank[1])})
        #print(reqarr['req'])
        
        
        
        #based on selected cable size, calculate all parameters to be copied to excel file
        resistance=df.at[reqarr['req'][1],'R']
        reactance=df.at[reqarr['req'][1],'X']
        noofruns=reqarr['req'][2]
    
        startingvoltagedrop=1.732*(startingcurrent/noofruns)*(length/1000)*(resistance*startingpowerfactor+reactance*sinPFstart)
        startingpercentagedrop="{:.2f}".format(startingvoltagedrop*100/voltage)
        
        runningvoltagedrop=1.732*(fullloadcurrent/noofruns)*(length/1000)*(resistance*runningpowerfactor+reactance*sinPF)
        runningpercentagedrop="{:.2f}".format(runningvoltagedrop*100/voltage)
    
        if maxmotor<=rating:
            deratingfactor=dfactor[0]
        else:
            deratingfactor=dfactor[1]
        
        
        ampacityofselectedcable=df.at[reqarr['req'][1],dfactor[2]]
        tst="{0}C x {1}".format(df.at[reqarr['req'][1],'Core'],df.at[reqarr['req'][1],'Size'])
        #print(tst)
        df1.loc[i,'Power Cable Size']=tst
        
        
        powerloss="{:.2f}".format(3*(fullloadcurrent**2)*resistance*length/1000000)
        #powerloss="{:.2f}".format((fullloadcurrent**2)*m.sqrt((resistance**2)+(reactance**2)))
        #print(fullloadcurrent**2,' ',m.sqrt((resistance**2)+(reactance**2)),' ',powerloss)
        df1.loc[i,'Power Loss']=powerloss
        powerlosspercent='{:.2f}'.format(float(powerloss)*100/rating)
        df1.loc[i,'Power Loss Percentage']=powerlosspercent
        df1.at[i,'No of runs']=noofruns
        df1.at[i,'Full Load Current']=fullloadcurrent
        df1.at[i,'Running VD']=runningpercentagedrop
        df1.at[i,'Starting VD']=startingpercentagedrop
        df1.at[i,'Starting Current']=startingcurrent
        df1.at[i,'Derating Factor']=deratingfactor
        #df1.at[i,'Min Cable Size for Fault current']=deratingfactor
        if df.at[reqarr['req'][1],'Conductor']=='AL':
            df1.at[i,'Min Cable Size for Fault current']=areaal
        else:
            df1.at[i,'Min Cable Size for Fault current']=areacu
        df1.at[i,'De-rated Cable Current Rating in Air']=float("{:.2f}". format(ampacityofselectedcable*deratingfactor)) 
        df1.at[i,'R']=resistance
        df1.at[i,'X']=reactance
        df1.at[i,'Cable Current Rating in Air']=ampacityofselectedcable
        df1.loc[i,'Type of Conductor']=df.at[reqarr['req'][1],'Conductor']
        df1.at[i,'Required Feeder Rating']=float(powerloss)+float(rating)
        #=rating+powerloss
        #print(rating)
        #print(powerloss)
        #print(rating+powerloss)
        
        if df.at[finalset[0][0][0],'Core']==1:
            reqdia={'reqd':(df.at[finalset[0][0][0],'Dia']*(finalset[0][1]*2-1)*2,finalset[0][0][0],finalset[0][1])}
            #print('single', reqdia, voltage, i)
            
        else:
            reqdia={'reqd':(df.at[finalset[0][0][0],'Dia']*finalset[0][1],finalset[0][0][0],finalset[0][1])}
            #print('multiple', reqdia, voltage, i)
            
        #check4- select cable size with minimum conductor size
        for rank1 in finalset:
            #print(rank)
            for inde1 in rank1[0]:
                if df.at[inde1,'Core']==1 and df.at[inde1,'Dia']*(rank1[1]*2-1)*2<reqdia['reqd'][0]:
                    reqdia.update({'reqd':(df.at[inde1,'Dia']*(rank1[1]*2-1)*2,inde1,rank1[1])})
                elif df.at[inde1,'Core']==3 and df.at[inde1,'Dia']*rank1[1]<reqdia['reqd'][0]:
                    reqdia.update({'reqd':(df.at[inde1,'Dia']*rank1[1],inde1,rank1[1])})
        
        #print(reqarr['req'][1]!=reqdia['reqd'][1])
        if reqarr['req'][1]!=reqdia['reqd'][1]:
            df1.loc[i,'Remark']='{0}R x {1}C x {2} {3} cable can also be selected based on laying area criteria.'.format(reqdia['reqd'][2],df.at[reqdia['reqd'][1],'Core'],df.at[reqdia['reqd'][1],'Size'],df.at[reqarr['req'][1],'Conductor'])
        #print(reqarr, voltage,i)
        formatoutmotor(faultcurrent,time,df.at[reqarr['req'][1],'Conductor'],df1.at[i,'Tag. No.'],rating,voltage,ampacityofselectedcable,deratingfactor,tst,PermissibleRunningVoltageDrop/100,PermissibleStartingVoltageDrop/100,resistance,reactance,noofruns,length,df.at[reqarr['req'][1],'Size'],runningpowerfactor,startingpowerfactor,motorefficiency)
    
        def calcGlands(dia,core,runs):
            global dbaddr
            dfg=pd.DataFrame()
            data=pd.read_excel(dbaddr+"Glands\\glandsData.xlsx", 'Baliga')
            dfg=dfg.append(data)
            indexGlands=dfg.index
            nrowsGlands=len(indexGlands)
        
        
            for iteratorGlands in range(1,nrowsGlands):
                if dfg.at[iteratorGlands,'ODMax'] >= dia >= dfg.at[iteratorGlands,'ODMin']:
                    selectedGlandSize = dfg.at[iteratorGlands,'Gland SizeMetric']
                    break
            else:
                selectedGlandSize = 'Data Not available'
                
            if int(core) == 3:
                quantity = runs * 2
            else:
                quantity = 3 * runs * 2
            
            return ([selectedGlandSize,quantity])       
        
        calculatedGlands = calcGlands(df.at[finalset[0][0][0],'Dia'], df.at[reqarr['req'][1],'Core'],noofruns)
        df1.loc[i,'Gland Size']=calculatedGlands[0]   
        df1.loc[i,'Glands Quantity']=calculatedGlands[1]
        df1.loc[i,'Dia']=df.at[finalset[0][0][0],'Dia']
        
        #print(reqarr['req'])
            
            #print(2, finalset)
        
        
    def HVfeeder(faultcurrent,time,rating,voltage,runningpowerfactor,dfactor,length):
        ty={}
        acceptable=[]
        acceptable1=[]
        acceptable2=[]
        finalset=[]
        #conductor='AL'
        global df1
        global i
        global nrows1
        global system
        global vendor
        global cbalum
        
        #importing cable data in dataframe
        df=pd.DataFrame()
        f=dbaddr+vendor+'\\'+'Cable data.xlsx'
        data=pd.read_excel(f,str(voltage)+system)
        df=df.append(data)
        index=df.index
        nrows=len(index)
        
        fullloadcurrent=float("{:.2f}". format(rating*1000/(1.732*voltage)))
        #startingcurrent=float("{:.2f}". format(fullloadcurrent*5*1.2))        
        
        #checkl- conductor size based on short circuit level
        
        
        sinPF=float("{:.2f}". format(m.sqrt(1-runningpowerfactor**2)))
        #sinPFstart=float("{:.2f}". format(m.sqrt(1-startingpowerfactor**2)))
        #conductor='Al'
        
        kal=0.094
        kcu=0.143
            
        areaal=float("{:.2f}". format(faultcurrent*m.sqrt(time)/kal))
        areacu=float("{:.2f}". format(faultcurrent*m.sqrt(time)/kcu))
        #print (cbalum)
        #checkl- conductor size based on short circuit level
        for k in range(1,nrows):
            if df.at[k,'Conductor']=='AL' and df.at[k,'Size']>=areaal:
                if df.at[k,'Core']==3 and df.at[k,'Size']<=maxcablesize:
                    acceptable.append(k)
                elif df.at[k,'Core']==1 and df.at[k,'Size']>maxcablesize:
                    acceptable.append(k)
                    
            elif df.at[k,'Conductor']=='CU' and df.at[k,'Size']>=areacu:
                if cbalum!='YES':             
                    if df.at[k,'Core']==3 and df.at[k,'Size']<=maxcablesize:
                        acceptable.append(k)
                    elif df.at[k,'Core']==1 and df.at[k,'Size']>maxcablesize:
                        acceptable.append(k)
        #print(acceptable, voltage)                
        for noofruns in range(1,6):
            
            
        #check2- filter cables based on ampacity    
            for l in acceptable:
                if df.at[l,'Core']==1:
                    deratingfactor=dfactor[0]
                else:
                    deratingfactor=dfactor[1]
                    
                ampacityofselectedcable=df.at[l,dfactor[2]]
                
                #deratingfactor=dfactor[1]
                designcurrent=float("{:.2f}". format(noofruns*ampacityofselectedcable*deratingfactor))        
                if designcurrent>=fullloadcurrent:
                    acceptable1.append(l)
            #print('acceptable1',acceptable1)      
            PermissibleRunningVoltageDrop = df1.at[i,'Permissible Running VD']
            
        #check3- filter cables based on running voltage drop    
            for p in acceptable1:
                #noofruns=m.ceil(fullloadcurrent/ty[p][0])
                resistance=df.at[p,'R']
                reactance=df.at[p,'X']
                runningvoltagedrop=1.732*(fullloadcurrent/noofruns)*(length/1000)*(resistance*runningpowerfactor+reactance*sinPF)
                #print(runningvoltagedrop/33)
                if runningvoltagedrop*100/voltage<=PermissibleRunningVoltageDrop:
                    acceptable2.append(p)
            #print('acceptable2',acceptable2)    
            PermissibleStartingVoltageDrop = df1.at[i,'Permissible Starting VD']
            
            
        #Create a tuple with all possible cable sizes for that rating
            try:
                if acceptable2!=[]:
                    finalset.append((acceptable2,noofruns))
            except:
                pass
            #finalset=list(set(finalset))
            #print('finalset',finalset)
            acceptable1=[]
            acceptable2=[]
        
        
        
        #initialize a selected cable size based on conductor material used
        reqarr={'req':(df.at[finalset[0][0][0],'Size']*finalset[0][1],finalset[0][0][0],finalset[0][1])}
        #print(reqarr)
        
        
        #check4- select cable size with minimum conductor size
        for rank in finalset:
            #print(rank)
            for inde in rank[0]:
                if df.at[inde,'Size']*rank[1]<reqarr['req'][0]:
                    reqarr.update({'req':(df.at[inde,'Size']*rank[1],inde,rank[1])})
        #print(reqarr['req'])
        
        
        
        #based on selected cable size, calculate all parameters to be copied to excel file
        resistance=df.at[reqarr['req'][1],'R']
        reactance=df.at[reqarr['req'][1],'X']
        noofruns=reqarr['req'][2]
    
        
        runningvoltagedrop=1.732*(fullloadcurrent/noofruns)*(length/1000)*(resistance*runningpowerfactor+reactance*sinPF)
        runningpercentagedrop="{:.2f}".format(runningvoltagedrop*100/voltage)
    
        if df.at[reqarr['req'][1],'Core']==1:
            deratingfactor=dfactor[0]
        else:
            deratingfactor=dfactor[1]
        
        
        ampacityofselectedcable=df.at[reqarr['req'][1],dfactor[2]]
        tst="{0}C x {1}".format(df.at[reqarr['req'][1],'Core'],df.at[reqarr['req'][1],'Size'])
        #print(tst)
        df1.loc[i,'Power Cable Size']=tst
        
        
        powerloss="{:.2f}".format(3*(fullloadcurrent**2)*resistance*length/1000000)
        #powerloss="{:.2f}".format((fullloadcurrent**2)*m.sqrt((resistance**2)+(reactance**2)))
        #print(fullloadcurrent**2,' ',m.sqrt((resistance**2)+(reactance**2)),' ',powerloss)
        df1.loc[i,'Power Loss']=powerloss
        df1.loc[i,'Power Loss Percentage']='{:.2f}'.format(float(powerloss)*100/rating)
        df1.at[i,'No of runs']=noofruns
        df1.at[i,'Full Load Current']=fullloadcurrent
        df1.at[i,'Running VD']=runningpercentagedrop
        
        df1.at[i,'Derating Factor']=deratingfactor
        #df1.at[i,'Min Cable Size for Fault current']=deratingfactor
        if df.at[reqarr['req'][1],'Conductor']=='AL':
            df1.at[i,'Min Cable Size for Fault current']=areaal
        else:
            df1.at[i,'Min Cable Size for Fault current']=areacu
        df1.at[i,'De-rated Cable Current Rating in Air']=float("{:.2f}". format(ampacityofselectedcable*deratingfactor)) 
        df1.at[i,'R']=resistance
        df1.at[i,'X']=reactance
        df1.at[i,'Cable Current Rating in Air']=ampacityofselectedcable
        df1.loc[i,'Type of Conductor']=df.at[reqarr['req'][1],'Conductor']
        df1.at[i,'Required Feeder Rating']=float(powerloss)+float(rating)
    
    
    
        
        if df.at[finalset[0][0][0],'Core']==1:
            reqdia={'reqd':(df.at[finalset[0][0][0],'Dia']*(finalset[0][1]*2-1)*2,finalset[0][0][0],finalset[0][1])}
            #print('single', reqdia, voltage, i)
            
        else:
            reqdia={'reqd':(df.at[finalset[0][0][0],'Dia']*finalset[0][1],finalset[0][0][0],finalset[0][1])}
            #print('multiple', reqdia, voltage, i)
            
        #check4- select cable size with minimum conductor size
        for rank1 in finalset:
            #print(rank)
            for inde1 in rank1[0]:
                if df.at[inde1,'Core']==1 and df.at[inde1,'Dia']*(rank1[1]*2-1)*2<reqdia['reqd'][0]:
                    reqdia.update({'reqd':(df.at[inde1,'Dia']*(rank1[1]*2-1)*2,inde1,rank1[1])})
                elif df.at[inde1,'Core']==3 and df.at[inde1,'Dia']*rank1[1]<reqdia['reqd'][0]:
                    reqdia.update({'reqd':(df.at[inde1,'Dia']*rank1[1],inde1,rank1[1])})
        
        #print(reqarr['req'][1]!=reqdia['reqd'][1])
        if reqarr['req'][1]!=reqdia['reqd'][1]:
            df1.loc[i,'Remark']='{0}R x {1}C x {2} {3} cable can also be selected based on laying area criteria.'.format(reqdia['reqd'][2],df.at[reqdia['reqd'][1],'Core'],df.at[reqdia['reqd'][1],'Size'],df.at[reqarr['req'][1],'Conductor'])
        #print(reqarr, voltage,i)
        #print(reqdia, voltage,i)
        formatoutfeeder(faultcurrent,time,df.at[reqarr['req'][1],'Conductor'],df1.at[i,'Tag. No.'],rating,voltage,ampacityofselectedcable,deratingfactor,tst,PermissibleRunningVoltageDrop/100,resistance,reactance,noofruns,length,df.at[reqarr['req'][1],'Size'],runningpowerfactor)

        def calcGlands(dia,core,runs):
            global dbaddr
            dfg=pd.DataFrame()
            data=pd.read_excel(dbaddr+"Glands\\glandsData.xlsx", 'Baliga')
            dfg=dfg.append(data)
            indexGlands=dfg.index
            nrowsGlands=len(indexGlands)
        
        
            for iteratorGlands in range(1,nrowsGlands):
                if dfg.at[iteratorGlands,'ODMax'] >= dia >= dfg.at[iteratorGlands,'ODMin']:
                    selectedGlandSize = dfg.at[iteratorGlands,'Gland SizeMetric']
                    break
            else:
                selectedGlandSize = 'Data Not available'
                
            if int(core) == 3:
                quantity = runs * 2
            else:
                quantity = 3 * runs * 2
            
            return ([selectedGlandSize,quantity])       
        
        calculatedGlands = calcGlands(df.at[finalset[0][0][0],'Dia'], df.at[reqarr['req'][1],'Core'],noofruns)
        df1.loc[i,'Gland Size']=calculatedGlands[0]   
        df1.loc[i,'Glands Quantity']=calculatedGlands[1]
        df1.loc[i,'Dia']=df.at[finalset[0][0][0],'Dia']
        
        
    def HVcapacitor(faultcurrent,time,rating,voltage,dfactor,length):
        ty={}
        acceptable=[]
        acceptable1=[]
        acceptable2=[]
        finalset=[]
        #conductor='AL'
        global df1
        global i
        global nrows1
        global system
        global vendor
        global cbalum
        #importing cable data in dataframe
        df=pd.DataFrame()
        f=dbaddr+vendor+'\\'+'Cable data.xlsx'
        data=pd.read_excel(f,str(voltage)+system)
        df=df.append(data)
        index=df.index
        nrows=len(index)
        
        fullloadcurrent=float("{:.2f}". format(rating*1000/(1.732*voltage)))
        #startingcurrent=float("{:.2f}". format(fullloadcurrent*5*1.2))        
        
        #checkl- conductor size based on short circuit level
        '''
        for k in range(1,nrows):
            if df.at[k,'Conductor']=='AL':
                conductor='AL'
                kf=0.094
            else:
                kf=0.143
                conductor='CU'
            area=float("{:.2f}". format(faultcurrent*m.sqrt(time)/kf))
            if df.at[k,'Size']>=area and df.at[k,'Conductor']==conductor:
                acceptable.append(k)
                
        #print(acceptable)
        '''
        
        runningpowerfactor=1
        sinPF=0
        #sinPF=float("{:.2f}". format(m.sqrt(1-runningpowerfactor**2)))
        #sinPFstart=float("{:.2f}". format(m.sqrt(1-startingpowerfactor**2)))
        #conductor='Al'
        
        kal=0.094
        kcu=0.143
            
        areaal=float("{:.2f}". format(faultcurrent*m.sqrt(time)/kal))
        areacu=float("{:.2f}". format(faultcurrent*m.sqrt(time)/kcu))

        
        #checkl- conductor size based on short circuit level
        #print(cbalum)
        for k in range(1,nrows):
            if df.at[k,'Conductor']=='AL' and df.at[k,'Size']>=areaal:
                if df.at[k,'Core']==3 and df.at[k,'Size']<=maxcablesize:
                    acceptable.append(k)
                elif df.at[k,'Core']==1 and df.at[k,'Size']>maxcablesize:
                    acceptable.append(k)
                    
            elif df.at[k,'Conductor']=='CU' and df.at[k,'Size']>=areacu:
                if cbalum!='YES':             
                    if df.at[k,'Core']==3 and df.at[k,'Size']<=maxcablesize:
                        acceptable.append(k)
                    elif df.at[k,'Core']==1 and df.at[k,'Size']>maxcablesize:
                        acceptable.append(k)
        #print('cap ', acceptable, voltage)
        for noofruns in range(1,6):
            
            
        #check2- filter cables based on ampacity    
            for l in acceptable:
                if df.at[l,'Core']==1:
                    deratingfactor=dfactor[0]
                else:
                    deratingfactor=dfactor[1]
                    
                ampacityofselectedcable=df.at[l,dfactor[2]]
                
                #deratingfactor=dfactor[1]
                designcurrent=float("{:.2f}". format(1.35*noofruns*ampacityofselectedcable*deratingfactor))        
                if designcurrent>=fullloadcurrent:
                    acceptable1.append(l)
            #print('acceptable1',acceptable1)      
            PermissibleRunningVoltageDrop = df1.at[i,'Permissible Running VD']
            
        #check3- filter cables based on running voltage drop    
            for p in acceptable1:
                #noofruns=m.ceil(fullloadcurrent/ty[p][0])
                resistance=df.at[p,'R']
                reactance=df.at[p,'X']
                runningvoltagedrop=1.732*(fullloadcurrent/noofruns)*(length/1000)*(resistance*runningpowerfactor+reactance*sinPF)
                #print(runningvoltagedrop/33)
                if runningvoltagedrop*100/voltage<=PermissibleRunningVoltageDrop:
                    acceptable2.append(p)
            #print('acceptable2',acceptable2)    
            PermissibleStartingVoltageDrop = df1.at[i,'Permissible Starting VD']
            
            
        #Create a tuple with all possible cable sizes for that rating
            try:
                if acceptable2!=[]:
                    finalset.append((acceptable2,noofruns))
            except:
                pass
            #finalset=list(set(finalset))
            #print('finalset',finalset)
            acceptable1=[]
            acceptable2=[]
        
        
        
        #initialize a selected cable size based on conductor material used
        reqarr={'req':(df.at[finalset[0][0][0],'Size']*finalset[0][1],finalset[0][0][0],finalset[0][1])}
        #print(reqarr)
        
        
        #check4- select cable size with minimum conductor size
        for rank in finalset:
            #print(rank)
            for inde in rank[0]:
                if df.at[inde,'Size']*rank[1]<reqarr['req'][0]:
                    reqarr.update({'req':(df.at[inde,'Size']*rank[1],inde,rank[1])})
        #print(reqarr['req'])
        
        
        
        #based on selected cable size, calculate all parameters to be copied to excel file
        resistance=df.at[reqarr['req'][1],'R']
        reactance=df.at[reqarr['req'][1],'X']
        noofruns=reqarr['req'][2]
    
        
        runningvoltagedrop=1.732*(fullloadcurrent/noofruns)*(length/1000)*(resistance*runningpowerfactor+reactance*sinPF)
        runningpercentagedrop="{:.2f}".format(runningvoltagedrop*100/voltage)
    
        if df.at[reqarr['req'][1],'Core']==1:
            deratingfactor=dfactor[0]
        else:
            deratingfactor=dfactor[1]
        
        
        ampacityofselectedcable=df.at[reqarr['req'][1],dfactor[2]]
        tst="{0}C x {1}".format(df.at[reqarr['req'][1],'Core'],df.at[reqarr['req'][1],'Size'])
        #print(tst)
        df1.loc[i,'Power Cable Size']=tst
        
        
        powerloss="{:.2f}".format(3*(fullloadcurrent**2)*resistance*length/1000000)
        #powerloss="{:.2f}".format((fullloadcurrent**2)*m.sqrt((resistance**2)+(reactance**2)))
        #print(fullloadcurrent**2,' ',m.sqrt((resistance**2)+(reactance**2)),' ',powerloss)
        df1.loc[i,'Power Loss']=powerloss
        df1.loc[i,'Power Loss Percentage']='{:.2f}'.format(float(powerloss)*100/rating)
        df1.at[i,'No of runs']=noofruns
        df1.at[i,'Full Load Current']=fullloadcurrent
        df1.at[i,'Running VD']=runningpercentagedrop
        
        df1.at[i,'Derating Factor']=deratingfactor
        #df1.at[i,'Min Cable Size for Fault current']=deratingfactor
        if df.at[reqarr['req'][1],'Conductor']=='AL':
            df1.at[i,'Min Cable Size for Fault current']=areaal
        else:
            df1.at[i,'Min Cable Size for Fault current']=areacu
        df1.at[i,'De-rated Cable Current Rating in Air']=float("{:.2f}". format(ampacityofselectedcable*deratingfactor)) 
        df1.at[i,'R']=resistance
        df1.at[i,'X']=reactance
        df1.at[i,'Cable Current Rating in Air']=ampacityofselectedcable
        df1.loc[i,'Type of Conductor']=df.at[reqarr['req'][1],'Conductor']
        df1.at[i,'Required Feeder Rating']=float(powerloss)+float(rating)    
    
        
        if df.at[finalset[0][0][0],'Core']==1:
            reqdia={'reqd':(df.at[finalset[0][0][0],'Dia']*(finalset[0][1]*2-1)*2,finalset[0][0][0],finalset[0][1])}
            #print('single', reqdia, voltage, i)
            
        else:
            reqdia={'reqd':(df.at[finalset[0][0][0],'Dia']*finalset[0][1],finalset[0][0][0],finalset[0][1])}
            #print('multiple', reqdia, voltage, i)
            
        #check4- select cable size with minimum conductor size
        for rank1 in finalset:
            #print(rank)
            for inde1 in rank1[0]:
                if df.at[inde1,'Core']==1 and df.at[inde1,'Dia']*(rank1[1]*2-1)*2<reqdia['reqd'][0]:
                    reqdia.update({'reqd':(df.at[inde1,'Dia']*(rank1[1]*2-1)*2,inde1,rank1[1])})
                elif df.at[inde1,'Core']==3 and df.at[inde1,'Dia']*rank1[1]<reqdia['reqd'][0]:
                    reqdia.update({'reqd':(df.at[inde1,'Dia']*rank1[1],inde1,rank1[1])})
        
        #print(reqarr['req'][1]!=reqdia['reqd'][1])
        if reqarr['req'][1]!=reqdia['reqd'][1]:
            df1.loc[i,'Remark']='{0}R x {1}C x {2} {3}cable can also be selected based on laying area criteria.'.format(reqdia['reqd'][2],df.at[reqdia['reqd'][1],'Core'],df.at[reqdia['reqd'][1],'Size'],df.at[reqarr['req'][1],'Conductor'])
        #print(reqarr, voltage,i)
        #print(reqdia, voltage,i)
        
        
        def calcGlands(dia,core,runs):
            global dbaddr
            dfg=pd.DataFrame()
            data=pd.read_excel(dbaddr+"Glands\\glandsData.xlsx", 'Baliga')
            dfg=dfg.append(data)
            indexGlands=dfg.index
            nrowsGlands=len(indexGlands)
        
        
            for iteratorGlands in range(1,nrowsGlands):
                if dfg.at[iteratorGlands,'ODMax'] >= dia >= dfg.at[iteratorGlands,'ODMin']:
                    selectedGlandSize = dfg.at[iteratorGlands,'Gland SizeMetric']
                    break
            else:
                selectedGlandSize = 'Data Not available'
                
            if int(core) == 3:
                quantity = runs * 2
            else:
                quantity = 3 * runs * 2
            
            return ([selectedGlandSize,quantity])       
        
        calculatedGlands = calcGlands(df.at[finalset[0][0][0],'Dia'], df.at[reqarr['req'][1],'Core'],noofruns)
        df1.loc[i,'Gland Size']=calculatedGlands[0]   
        df1.loc[i,'Glands Quantity']=calculatedGlands[1]
        df1.loc[i,'Dia']=df.at[finalset[0][0][0],'Dia']
            
    #Dsa=0.81
    #Dma=0.63
    #Dsg=1
    #Dmg=1
        
    def calc_cables1():    
        global nrows1
        global df1
        global i
        global f1
        global calccablesize1
        global faddr
        global Dsa1
        global Dma1
        global Dsg1
        global Dmg1
        global vendor
        df3 = pd.DataFrame()
        #f1=faddr+'Test.xlsx'
        #f=r'C:\Users\20315914\OneDrive - Larsen & Toubro\Desktop\Test1.xlsx'
        data1 = pd.read_excel(f1, 'HV') 
        df3 = df3.append(data1)
        df1=df3.iloc[0:,[i for i in range(30)]+[38]]
        df4=df3.iloc[0:,[30,31,32,33,34,35,36,37]]
        index1=df1.index
        nrows1=len(index1)    
        df1.insert(30,'Power Loss','')
        df1.loc[0,'Power Loss']='(kW)'
        df1.insert(31,'Power Loss Percentage','')
        df1.loc[0,'Power Loss Percentage']='(%)'
        df1.insert(32,'Required Feeder Rating','')
        df1.loc[0,'Required Feeder Rating']='(kW)'
        df1.insert(33,'Dia','')
        df1.loc[0,'Dia']='(mm)'
        df1.insert(34,'Gland Size','')
        df1.loc[0,'Gland Size']='(Metric)'
        df1.insert(35,'Glands Quantity','')
        df1.loc[0,'Glands Quantity']='(Nos.)'
        for i in range(1,nrows1):
                
            if df1.at[i,'Feeder type']=='Power':
                
                try:
                    if df1.loc[i,'Cable laying']=='AG':
                        dfactor=[Dsa1,Dma1,'Ia']
                    elif df1.loc[i,'Cable laying']=='UG':
                        dfactor=[Dsg1,Dmg1,'Ig']
                    else:
                        raise Exception
                    #pass
                    #listtype=[df1.at[i,'Fault Current'],df1.at[i,'Fault Clearing Time'],df1.at[i,'Selected Power Rating'],df1.at[i,'Voltage'],df1.at[i,'PF'],dfactor,df1.at[i,'Cable Length']]
                    #print(listtype)
                    
                    #print(1)
                    HVfeeder(df1.at[i,'Fault Current'],df1.at[i,'Fault Clearing Time'],df1.at[i,'Selected Power Rating'],df1.at[i,'Voltage'],df1.at[i,'PF'],dfactor,df1.at[i,'Cable Length'])
                except:
                    df1.loc[i,'Remark']='Incomplete Data Provided'
                    #print('test1')
            elif df1.at[i,'Feeder type']=='Motor':
                
                try:
                    if df1.loc[i,'Cable laying']=='AG':
                        dfactor=[Dsa1,Dma1,'Ia']
                    elif df1.loc[i,'Cable laying']=='UG':
                        dfactor=[Dsg1,Dmg1,'Ig']
                    else:
                        raise Exception
                    #listtype=[df1.at[i,'Fault Current'],df1.at[i,'Fault Clearing Time'],df1.at[i,'Absorbed Power'],df1.at[i,'Voltage'],df1.at[i,'PF'],df1.at[i,'Starting PF'],df1.at[i,'Efficiency'],dfactor,df1.at[i,'Cable Length']]
                    #print(listtype)
                    
                    HVmotor(df1.at[i,'Fault Current'],df1.at[i,'Fault Clearing Time'],df1.at[i,'Absorbed Power'],df1.at[i,'Voltage'],df1.at[i,'PF'],df1.at[i,'Starting PF'],df1.at[i,'Efficiency'],dfactor,df1.at[i,'Cable Length'],df1.at[i,'Motor Starting Method'])
                except:
                    df1.loc[i,'Remark']='Incomplete Data Provided'    
                    #print('test2')
            elif df1.at[i,'Feeder type']=='Capacitor':                
                
                try:
                    if df1.loc[i,'Cable laying']=='AG':
                        dfactor=[Dsa1,Dma1,'Ia']
                    elif df1.loc[i,'Cable laying']=='UG':
                        dfactor=[Dsg1,Dmg1,'Ig']
                    else:
                        raise Exception
                    #listtype=[df1.at[i,'Fault Current'],df1.at[i,'Fault Clearing Time'],df1.at[i,'Selected Power Rating'],df1.at[i,'Voltage'],dfactor,df1.at[i,'Cable Length']]
                    #print(listtype)
                    
                    HVcapacitor(df1.at[i,'Fault Current'],df1.at[i,'Fault Clearing Time'],df1.at[i,'Selected Power Rating'],df1.at[i,'Voltage'],dfactor,df1.at[i,'Cable Length'])
                    #pass
                except:
                    df1.loc[i,'Remark']='Incomplete Data Provided' 
                    #print('test3')
            else:
                df1.loc[i,'Remark']='Incomplete Data Provided' 
                
        writer=pd.ExcelWriter(f2+'/output/HV_Cable Sizing_{}.xlsx'.format(vendor),engine='xlsxwriter')
        workbook  = writer.book
        df1.to_excel(writer, sheet_name='Sheet1',index=False,na_rep='-')
        worksheet = writer.sheets['Sheet1']
        border_format=workbook.add_format({'border':1,'align':'centre','font_size':10})
        end='AH'+str(nrows1+1)
        worksheet.conditional_format( 'A1:{0}'.format(end) , { 'type' : 'no_blanks' , 'format' : border_format} )
        #worksheet.set_column('A:D',12,border_format)    
        df2=myfunc(df1)
        df2.to_excel(writer, sheet_name='HV Cable Lot-1',index=False)
        worksheet = writer.sheets['HV Cable Lot-1']
        border_format=workbook.add_format({'border':1,'align':'centre','font_size':10})
        end='AH'+str(nrows1+1)
        worksheet.conditional_format( 'A1:{0}'.format(end) , { 'type' : 'no_blanks' , 'format' : border_format} )
        

        df5=cableGlands(df1)
        df5.to_excel(writer, sheet_name='Cable Gland Lot-1',index=False)
        worksheet = writer.sheets['Cable Gland Lot-1']
        border_format=workbook.add_format({'border':1,'align':'centre','font_size':10})
        end='AH'+str(nrows1+1)
        worksheet.conditional_format( 'A1:{0}'.format(end) , { 'type' : 'no_blanks' , 'format' : border_format} )
    
        
        df4.to_excel(writer, sheet_name='Transformer Details',index=False)
        worksheet = writer.sheets['Transformer Details']
        border_format=workbook.add_format({'border':1,'align':'centre','font_size':10})
        end='AH'+str(nrows1+1)
        worksheet.conditional_format( 'A1:{0}'.format(end) , { 'type' : 'no_blanks' , 'format' : border_format} )
        
        workbook.close()
        design('HV_Cable Sizing_{}'.format(vendor))
        calccablesize1['state']=DISABLED
        messagebox.showinfo('SUCCESS','Cable Sizing Completed Successfully')
        #df1.to_excel(faddr+'HV.xlsx',index=False) #resultant data is stored in this excel sheet
        #calccablesize['state']=DISABLED
        #data = pd.read_excel(r'C:\Users\20315914\OneDrive - Larsen & Toubro\Desktop\Test2.xlsx') 
 

        '''
        def HVfeeder(faultcurrent,time,conductor,rating,voltage,runningpowerfactor,deratingfactor,length):
            # short circuit withstand capacity of cable
            #faultcurrent=40
            #time=1
            global df1
            global j
            if conductor=='AL':
                k=0.094
            else:
                k=0.143
            area=faultcurrent*m.sqrt(time)/k
            #print ("{:.2f}". format(area))
            
            #Cable ampacity for load current
            #rating=9215
            #voltage=11000
            #runningpowerfactor=0.8
            fullloadcurrent=float("{:.2f}". format(rating*1000/(1.732*voltage)))
            sinPF=float("{:.2f}". format(m.sqrt(1-runningpowerfactor**2)))
            #deratingfactor=0.81
            ampacityofselectedcable=840
            designcurrent=float("{:.2f}". format(ampacityofselectedcable*deratingfactor))
            #print(designcurrent)
            noofruns=m.ceil(fullloadcurrent/designcurrent)
            #print(noofruns)
            
            #voltage drop consideration
            PermissibleRunningVoltageDrop = 0.8
            resistance=0.06
            reactance=0.094
            #length=980
            
            runningvoltagedrop=1.732*(fullloadcurrent/noofruns)*(length/1000)*(resistance*runningpowerfactor+reactance*sinPF)
            percentagedrop=float("{:.3f}". format((runningvoltagedrop/voltage)*100))
            df1.at[j,'Z']=fullloadcurrent
            df1.at[j,'AA']=percentagedrop
            #df.at(i,'AB')=startingpercentagedrop
            #print(percentagedrop)
        
        def LVfeeder(faultcurrent,time,conductor,rating,voltage,runningpowerfactor,deratingfactor,length):
            # short circuit withstand capacity of cable
            #faultcurrent=40
            #time=1
            global df
            global i
            if conductor=='AL':
                k=0.094
            else:
                k=0.143
            area=faultcurrent*m.sqrt(time)/k
            #print ("{:.2f}". format(area))
            
            #Cable ampacity for load current
            #rating=9215
            #voltage=11000
            #runningpowerfactor=0.8
            fullloadcurrent=float("{:.2f}". format(rating*1000/(1.732*voltage)))
            sinPF=float("{:.2f}". format(m.sqrt(1-runningpowerfactor**2)))
            #deratingfactor=0.81
            ampacityofselectedcable=840
            designcurrent=float("{:.2f}". format(ampacityofselectedcable*deratingfactor))
            #print(designcurrent)
            noofruns=m.ceil(fullloadcurrent/designcurrent)
            #print(noofruns)
            
            #voltage drop consideration
            PermissibleRunningVoltageDrop = 0.8
            resistance=0.06
            reactance=0.094
            #length=980
            
            runningvoltagedrop=1.732*(fullloadcurrent/noofruns)*(length/1000)*(resistance*runningpowerfactor+reactance*sinPF)
            percentagedrop=float("{:.3f}". format((runningvoltagedrop/voltage)*100))
            df.at[i,'Z']=fullloadcurrent
            df.at[i,'AA']=percentagedrop
            #df.at(i,'AB')=startingpercentagedrop
            #print(percentagedrop)
        
        
        def LVPfeeder(rating,voltage,runningpowerfactor,deratingfactor,length):
            # short circuit withstand capacity of cable
            #faultcurrent=40
            #time=1
            global df
            global i
            
            fullloadcurrent=float("{:.2f}". format(rating*1000/(1.732*voltage)))
            sinPF=float("{:.2f}". format(m.sqrt(1-runningpowerfactor**2)))
            
            ampacityofselectedcable=840
            designcurrent=float("{:.2f}". format(ampacityofselectedcable*deratingfactor))
            
            noofruns=m.ceil(fullloadcurrent/designcurrent)
            
            
            PermissibleRunningVoltageDrop = 0.8
            resistance=0.06
            reactance=0.094
            #length=980
            
            runningvoltagedrop=1.732*(fullloadcurrent/noofruns)*(length/1000)*(resistance*runningpowerfactor+reactance*sinPF)
            percentagedrop=float("{:.3f}". format((runningvoltagedrop/voltage)*100))
            df.at[i,'Z']=fullloadcurrent
            df.at[i,'AA']=percentagedrop
            #df.at(i,'AB')=startingpercentagedrop
            #print(percentagedrop)
        
        def HVmotor(faultcurrent,time,conductor,rating,voltage,runningpowerfactor,startingpowerfactor,motorefficiency,deratingfactor,length):
            # short circuit withstand capacity of cable
            #faultcurrent=40
            #time=1
            global df1
            global j
            
            if conductor=='AL':
                k=0.094
            else:
                k=0.143
            area=faultcurrent*m.sqrt(time)/k
            #print ("{:.2f}". format(area))
            
            #Cable ampacity for load current
            #rating=9215
            #voltage=11000
            #runningpowerfactor=0.8
            fullloadcurrent=float("{:.2f}". format(rating*1000/(1.732*voltage*runningpowerfactor*motorefficiency)))
            startingcurrent=float("{:.2f}". format(fullloadcurrent*5*1.2))
            #print(fullloadcurrent,startingcurrent)
            sinPF=float("{:.2f}". format(m.sqrt(1-runningpowerfactor**2)))
            sinPFstart=float("{:.2f}". format(m.sqrt(1-startingpowerfactor**2)))
            #print(sinPFstart)
            #deratingfactor=0.81
            ampacityofselectedcable=295
            designcurrent=float("{:.2f}". format(ampacityofselectedcable*deratingfactor))
            #print(designcurrent)
            noofruns=m.ceil(fullloadcurrent/designcurrent)
            #print(noofruns)
            
            #voltage drop consideration
            PermissibleRunningVoltageDrop = 0.8
            resistance=0.264
            reactance=0.076
            #length=980
            
            runningvoltagedrop=1.732*(fullloadcurrent/noofruns)*(length/1000)*(resistance*runningpowerfactor+reactance*sinPF)
            runningpercentagedrop=float("{:.3f}". format((runningvoltagedrop/voltage)*100))
            #print(runningvoltagedrop,runningpercentagedrop)
            startingvoltagedrop=1.732*(startingcurrent/noofruns)*(length/1000)*(resistance*startingpowerfactor+reactance*sinPFstart)
            startingpercentagedrop=float("{:.3f}". format((startingvoltagedrop/voltage)*100))
            df1.at[j,'Z']=fullloadcurrent
            df1.at[j,'AA']=runningpercentagedrop
            df1.at[j,'AB']=startingpercentagedrop
            
            #print(startingvoltagedrop,startingpercentagedrop)
        
        
        
        def LVmotor(faultcurrent,time,conductor,rating,voltage,runningpowerfactor,startingpowerfactor,motorefficiency,deratingfactor,length):
            # short circuit withstand capacity of cable
            #faultcurrent=40
            #time=1
            global df
            global i
            
            if conductor=='AL':
                k=0.094
            else:
                k=0.143
            area=faultcurrent*m.sqrt(time)/k
            #print ("{:.2f}". format(area))
            
            #Cable ampacity for load current
            #rating=9215
            #voltage=11000
            #runningpowerfactor=0.8
            fullloadcurrent=float("{:.2f}". format(rating*1000/(1.732*voltage*runningpowerfactor*motorefficiency)))
            startingcurrent=float("{:.2f}". format(fullloadcurrent*5*1.2))
            #print(fullloadcurrent,startingcurrent)
            sinPF=float("{:.2f}". format(m.sqrt(1-runningpowerfactor**2)))
            sinPFstart=float("{:.2f}". format(m.sqrt(1-startingpowerfactor**2)))
            #print(sinPFstart)
            #deratingfactor=0.81
            ampacityofselectedcable=295
            designcurrent=float("{:.2f}". format(ampacityofselectedcable*deratingfactor))
            #print(designcurrent)
            noofruns=m.ceil(fullloadcurrent/designcurrent)
            #print(noofruns)
            
            #voltage drop consideration
            PermissibleRunningVoltageDrop = 0.8
            resistance=0.264
            reactance=0.076
            #length=980
            
            runningvoltagedrop=1.732*(fullloadcurrent/noofruns)*(length/1000)*(resistance*runningpowerfactor+reactance*sinPF)
            runningpercentagedrop=float("{:.3f}". format((runningvoltagedrop/voltage)*100))
            #print(runningvoltagedrop,runningpercentagedrop)
            startingvoltagedrop=1.732*(startingcurrent/noofruns)*(length/1000)*(resistance*startingpowerfactor+reactance*sinPFstart)
            startingpercentagedrop=float("{:.3f}". format((startingvoltagedrop/voltage)*100))
            df.at[i,'Z']=fullloadcurrent
            df.at[i,'AA']=runningpercentagedrop
            df.at[i,'AB']=startingpercentagedrop
            
            #print(startingvoltagedrop,startingpercentagedrop)
'''    
    
    def getExcel ():
        pass
        
    
    
    def calc_cables():    
        pass   
    
    def back1():
        global cablesizing
        HVcablesize['state']=NORMAL
        LVcablesize['state']=NORMAL
        cableschedule1['state']=NORMAL
        drumschedule1['state']=NORMAL
        cablesizing.destroy()
    
    
    def back2():
        global cablesizing
        LVcablesize['state']=NORMAL
        HVcablesize['state']=NORMAL
        cableschedule1['state']=NORMAL
        drumschedule1['state']=NORMAL
        cablesizing.destroy()
    
    def back3():
        global cableschedule
        LVcablesize['state']=NORMAL
        
        cableschedule1['state']=NORMAL
        drumschedule1['state']=NORMAL
        cableschedule.destroy()
        
    def back4():
        global drumschedule
        LVcablesize['state']=NORMAL
        
        cableschedule1['state']=NORMAL
        drumschedule1['state']=NORMAL
        drumschedule.destroy()
    
    
    
    
    def LVcable_sizing():
        global nrows
        global df
        global f
        global back
        global cablesizing
        global mainwindow
        #global INPUT1
        #global INPUT2
        #global INPUT3
        #global INPUT4
        #global INPUT5
        #global INPUT6
        global INPUT7
        global INPUT8
        global INPUT9
        global INPUT10
        global INPUT11
        global INPUT12
        #global reset
        global reset1
        global browseButton_Excel1
        #global browseButton_Excel
        #global calccablesize
        global calccablesize1
        #global var
        #global var1
        global dbaddr
        dbaddr = askdirectory(title='SELECT DATABASE FOLDER')+'/'
        print(dbaddr)
        def disable_event():
            messagebox.showinfo('ERROR','KINDLY USE EXIT BUTTON')
        
        systeme1=['UN-EARTHED SYSTEM','EARTHED SYSTEM']

        #var = IntVar()
        #var1 = IntVar()
        mainwindow.iconify()
        
        
        HVcablesize['state']=DISABLED
        cableschedule1['state']=DISABLED
        drumschedule1['state']=DISABLED
        LVcablesize['state']=DISABLED
        cablesizing=Tk()
        cablesizing.title('HV CABLE SIZING')
        cablesizing.geometry('1390x680+5+5')
        cablesizing.configure(bg='steelblue')
        back=Button(cablesizing, width=10,text='EXIT',fg='blue',font=('Times New Roman', 16, 'bold'),command=back2)
        back.place(x=732,y=620)
        cablesizing.protocol("WM_DELETE_WINDOW", disable_event)
        global cb1
        cb1=Combobox(cablesizing,width=22, values=systeme1,font=('Times New Roman',12,'bold'))
        cb1.current(0)
        cb1.place(x=1050,y=138)
        #calccablesize=Button(cablesizing, width=10,text='RUN SIZING',fg='blue',font=('Times New Roman', 16, 'bold'),command=calc_cables)
        #calccablesize.place(x=395,y=620)
        calccablesize1=Button(cablesizing, width=10,text='RUN SIZING',fg='blue4',font=('Times New Roman', 16, 'bold'),command=calc_cables1)
        calccablesize1.place(x=307,y=620)
        heading=Label(cablesizing,font=('arial', 30,'bold'),text='HV CABLE SIZING',fg='blue')
        heading.pack()
        #scopeLV=Label(cablesizing,font=('arial', 30,'bold'),text='LV',fg='blue')
        #scopeLV.place(x=100, y=10)
        #scopeHV=Label(cablesizing,font=('arial', 30,'bold'),text='HV',fg='blue')
        #scopeHV.place(x=1180, y=10)
        #R1 = Radiobutton(cablesizing, fg='red', text="Earthed System", variable=var, value=2,command=sel)
        #R1.place(x=180,y=20)
        #R2 = Radiobutton(cablesizing,fg='red', text="Un-Earthed System", variable=var, value=1,command=sel1)
        #R2.place(x=300,y=20)
        #R3 = Radiobutton(cablesizing, fg='red',text="Earthed System", variable=var1, value=2,command=sel2)
        #R3.place(x=910,y=20)
        #R4 = Radiobutton(cablesizing, fg='red',text="Un-Earthed System", variable=var1, value=1,command=sel3)
        #R4.place(x=1030,y=20)
        
        #vendorlabel1=Label(cablesizing,fg='red',text='PLEASE SELECT PREFERRED VENDOR')
        #vendorlabel1.place(x=900, y=70)
        
        def Take_input15():
            global maxmotor
            try:
                maxmotor = float(scmotorrating1.get("1.0", "end-1c"))
                #print(type(INPUT7))
                #print(maxmotor)
            except(Exception):
                
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
        
        scoremotor1=Label(cablesizing,fg='red',text='PLEASE PROVIDE MOTOR RATING IN KW \n ABOVE WHICH SINGLE CORE CABLE IS TO BE USED \n DEFAULT RATING IS 2MW')
        scoremotor1.place(x=750, y=190)
        scmotorrating1 = Text(cablesizing, height = 1,width = 5,bg = "light yellow")
        scmotorrating1.place(x=1050,y=190)   
        go15 = Button(cablesizing, height = 1,width = 8,text ="CHECK",command = Take_input15)
        go15.place(x=1115,y=190) 
        #ratingl=Label(cablesizing,fg='red',text='DEFAULT =2MW')
        #ratingl.place(x=1200, y=240)
        
        def del_input15():
            global maxmotor            
            maxmotor=2000   
            scmotorrating1.delete("1.0", "end")
            #print(maxmotor)
        go151 = Button(cablesizing, height = 1,width = 8,text ="RESET",command = del_input15)
        go151.place(x=1190,y=190)
        
        def Take_input16():
            global maxcablesize
            try:
                maxcablesize = float(maxcablesize1.get("1.0", "end-1c"))
                #print(type(INPUT7))
                #print(maxcablesize)
            except(Exception):
                
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
        
        maxcablesize11=Label(cablesizing,fg='red',text='PLEASE PROVIDE CABLE SIZE IN SQ.MM. \n ABOVE WHICH SINGLE CORE CABLE IS TO BE USED \n DEFAULT SIZE IS 300 SQ.MM.')
        maxcablesize11.place(x=750, y=260)
        maxcablesize1 = Text(cablesizing, height = 1,width = 5,bg = "light yellow")
        maxcablesize1.place(x=1050,y=260)   
        go16 = Button(cablesizing, height = 1,width = 8,text ="CHECK",command = Take_input16)
        go16.place(x=1115,y=260) 
        #maxsize1=Label(cablesizing,fg='red',text='DEFAULT =300 SQ.MM.')
        #maxsize1.place(x=1250, y=340)
        
        def del_input16():
            global maxcablesize
            maxcablesize=300
            maxcablesize1.delete("1.0", "end")
            #print(maxcablesize)
        
        go161 = Button(cablesizing, height = 1,width = 8,text ="RESET",command = del_input16)
        go161.place(x=1190,y=260) 
        
        alumdecision=Label(cablesizing,fg='red',text='DO YOU WANT ALUMINIUM CABLES ONLY?')
        alumdecision.place(x=750, y=330)   
        cbalum1=Combobox(cablesizing,width=4, values=['YES','NO'],font=('Times New Roman',12,'bold'))
        cbalum1.current(0)
        cbalum1.place(x=1050,y=330)
        cbalum=cbalum1.get()
        
        scurrent=Label(cablesizing,fg='red',text='PLEASE PROVIDE STARTING CURRENT AS MULTIPLE OF FLC, 20% TOLERANCE WILL BE ADDED INTERNALLY')
        scurrent.place(x=750, y=370)
        
        def Take_input17():
            global factor1
            try:
                factor1 = float(DOL.get("1.0", "end-1c"))
                #print(type(INPUT7))
                #print(maxcablesize)
            except(Exception):
                
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
        
        DOLlabel=Label(cablesizing,fg='red',text='DOL \n DEFAULT=5 TIMES')
        DOLlabel.place(x=750, y=410)
        DOL = Text(cablesizing, height = 1,width = 5,bg = "light yellow")
        DOL.place(x=750,y=453)   
        go17 = Button(cablesizing, height = 1,width = 8,text ="CHECK",command = Take_input17)
        go17.place(x=750,y=480) 
        #maxsize1=Label(cablesizing,fg='red',text='DEFAULT =300 SQ.MM.')
        #maxsize1.place(x=1250, y=340)
        
        def del_input17():
            global factor1
            factor1=5
            DOL.delete("1.0", "end")
            #print(maxcablesize)
        
        go171 = Button(cablesizing, height = 1,width = 8,text ="RESET",command = del_input17)
        go171.place(x=750,y=510)
        
        def Take_input18():
            global factor2
            try:
                factor2 = float(VFD.get("1.0", "end-1c"))
                #print(type(INPUT7))
                #print(maxcablesize)
            except(Exception):
                
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
        
        VFDlabel=Label(cablesizing,fg='red',text='VFD \n DEFAULT=1.5 TIMES')
        VFDlabel.place(x=900, y=410)
        VFD = Text(cablesizing, height = 1,width = 5,bg = "light yellow")
        VFD.place(x=900,y=453)   
        go18 = Button(cablesizing, height = 1,width = 8,text ="CHECK",command = Take_input18)
        go18.place(x=900,y=480) 
        #maxsize1=Label(cablesizing,fg='red',text='DEFAULT =300 SQ.MM.')
        #maxsize1.place(x=1250, y=340)
        
        def del_input18():
            global factor2
            factor2=1.5
            VFD.delete("1.0", "end")
            #print(maxcablesize)
        
        go181 = Button(cablesizing, height = 1,width = 8,text ="RESET",command = del_input18)
        go181.place(x=900,y=510)
        
        
        def Take_input19():
            global factor3
            try:
                factor3 = float(SS.get("1.0", "end-1c"))
                #print(type(INPUT7))
                #print(maxcablesize)
            except(Exception):
                
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
        
        SSlabel=Label(cablesizing,fg='red',text='SOFT STARTER \n DEFAULT=1 TIMES')
        SSlabel.place(x=1050, y=410)
        SS = Text(cablesizing, height = 1,width = 5,bg = "light yellow")
        SS.place(x=1050,y=453)   
        go19 = Button(cablesizing, height = 1,width = 8,text ="CHECK",command = Take_input19)
        go19.place(x=1050,y=480) 
        #maxsize1=Label(cablesizing,fg='red',text='DEFAULT =300 SQ.MM.')
        #maxsize1.place(x=1250, y=340)
        
        def del_input19():
            global factor3
            factor3=1
            SS.delete("1.0", "end")
            #print(maxcablesize)
        
        go191 = Button(cablesizing, height = 1,width = 8,text ="RESET",command = del_input19)
        go191.place(x=1050,y=510)
        
        
        def Take_input20():
            global factor4
            try:
                factor4 = float(SD.get("1.0", "end-1c"))
                #print(type(INPUT7))
                #print(maxcablesize)
            except(Exception):
                
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
        
        SDlabel=Label(cablesizing,fg='red',text='STAR DELTA \n DEFAULT=3 TIMES')
        SDlabel.place(x=1200, y=410)
        SD = Text(cablesizing, height = 1,width = 5,bg = "light yellow")
        SD.place(x=1200,y=453)   
        go20 = Button(cablesizing, height = 1,width = 8,text ="CHECK",command = Take_input20)
        go20.place(x=1200,y=480) 
        #maxsize1=Label(cablesizing,fg='red',text='DEFAULT =300 SQ.MM.')
        #maxsize1.place(x=1250, y=340)
        
        def del_input20():
            global factor4
            factor4=3
            SD.delete("1.0", "end")
            #print(maxcablesize)
        
        go201 = Button(cablesizing, height = 1,width = 8,text ="RESET",command = del_input20)
        go201.place(x=1200,y=510)

        w = Canvas(cablesizing, width=2, height=500)
        w.create_line(700, 100, 2, 500)
        w.place(x=700,y=80)
        
        
        
        cb7=Combobox(cablesizing,width=2, values=['A','M'],font=('Times New Roman',12,'bold'))
        cb7.current(0)
        cb7.place(x=570,y=67)
        def Take_input7():
            global INPUT7
            global cb7a
            try:
                if cb7.get()=='M':
                    INPUT7 = float(hvinputatemp.get("1.0", "end-1c"))
                    if 1.14<INPUT7 or INPUT7<0.9:
        
                        raise Exception
                    #print(type(INPUT7))
                else:
                    INPUT7=float(cb7a.get())
            except(Exception):
                INPUT7=1
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
                
        hvatempderatingfactor = Label(cablesizing,fg='red',text = 'Please Enter Temperature Derating Factor (AG)= Kta')
        hvatempderatingfactor.place(x=100,y=70)    
        hvinputatemp = Text(cablesizing, height = 1.3,width = 5,bg = "light yellow")
        hvinputatemp.place(x=420,y=70)    
    
        go7 = Button(cablesizing, height = 1,width = 10,text ="CHECK",command = Take_input7)
        go7.place(x=480,y=67)  
        
        
          
        
        def info7display():
            global dbaddr
            global vendor
            img = cv.imread(dbaddr+vendor+'\\'+"Temperature derating factor_HV air.jpg",1)
            #print(img)
            #img = cv.resize(img, (850, 670))
            cv.imshow("Temperature Derating Factor in Air", img)
            
            cv.waitKey(delay=0)
            cv.destroyAllWindows()
                      
            
        info7 = Button(cablesizing, height = 1,width = 10,text ="INFO",command = info7display)
        info7.place(x=12,y=67)    
        info7['state']=DISABLED
        
        cb8=Combobox(cablesizing,width=2, values=['A','M'],font=('Times New Roman',12,'bold'))
        cb8.current(0)
        cb8.place(x=570,y=117)
        def Take_input8():
            global INPUT8
            #global cb8
            global cb8a
            try:
                if cb8.get()=='M':
                    
                        INPUT8 = float(hvinputgtemp.get("1.0", "end-1c"))
                        #print(type(INPUT8))
                        if 1.12<INPUT8 or INPUT8<0.87:
                            
                            raise Exception
                else:
                    INPUT8=float(cb8a.get())
                    #print('yyay')
            except(Exception):
                INPUT8=1
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
                #print(Exception)
                
        hvgtempderatingfactor = Label(cablesizing,fg='red',text = 'Please Enter Temperature Derating Factor (UG)= Ktg')
        hvgtempderatingfactor.place(x=100,y=120)    
        hvinputgtemp = Text(cablesizing, height = 1.3,width = 5,bg = "light yellow")
        hvinputgtemp.place(x=420,y=120)    
    
        go8 = Button(cablesizing, height = 1,width = 10,text ="CHECK",command = Take_input8)
        go8.place(x=480,y=117) 
        
        
        def info8display():
            global dbaddr
            global vendor
            img = cv.imread(dbaddr+vendor+'\\'+"Temperature derating factor_HV ground.jpg",1)
            #print(img)
            #img = cv.resize(img, (850, 670))
            cv.imshow("Temperature Derating Factor in Ground", img)
            
            cv.waitKey(delay=0)
            cv.destroyAllWindows()
                 
        info8 = Button(cablesizing, height = 1,width = 10,text ="INFO",command = info8display)
        info8.place(x=12,y=117)  
        info8['state']=DISABLED
        
        cb9=Combobox(cablesizing,width=2, values=['A','M'],font=('Times New Roman',12,'bold'))
        cb9.current(0)
        cb9.place(x=570,y=167)
        def Take_input9():
            global INPUT9
            global cb9a
            try:
                if cb9.get()=='M':
                    INPUT9 = float(hvinputdepth.get("1.0", "end-1c"))
                    #print(type(INPUT9))
                    if 1<INPUT9 or INPUT9<0.96:
                        
                        raise Exception
                else:
                    INPUT9=float(cb9a.get())
            except(Exception):
                INPUT9=1
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
                
        hvdepthderatingfactor = Label(cablesizing,fg='red',text = 'Please Enter Depth of Laying Derating Factor= Kd')
        hvdepthderatingfactor.place(x=100,y=170)    
        hvinputdepth = Text(cablesizing, height = 1.3,width = 5,bg = "light yellow")
        hvinputdepth.place(x=420,y=170)    
    
        go9 = Button(cablesizing, height = 1,width = 10,text ="CHECK",command = Take_input9)
        go9.place(x=480,y=167)    
        
        
        def info9display():
            global dbaddr
            global vendor
            img = cv.imread(dbaddr+vendor+'\\'+"Laying depth_HV.jpg",1)
            #print(img)
            #img = cv.resize(img, (950, 690))
            cv.imshow("Laying Depth Derating Factor for HV Cables", img)
            
            cv.waitKey(delay=0)
            cv.destroyAllWindows()
                 
        info9 = Button(cablesizing, height = 1,width = 10,text ="INFO",command = info9display)
        info9.place(x=12,y=167) 
        info9['state']=DISABLED
        
        
        cb10=Combobox(cablesizing,width=2, values=['A','M'],font=('Times New Roman',12,'bold'))
        cb10.current(0)
        cb10.place(x=570,y=217)
        def Take_input10():
            global INPUT10
            global cb10a
            try:
                if cb10.get()=='M':
                    INPUT10 = float(hvinputthermalresis.get("1.0", "end-1c"))
                    if 1.2<INPUT10 or INPUT10<0.73:
                        
                        raise Exception
                    #print(type(INPUT10))
                else:
                    INPUT10=float(cb10a.get())
            except(Exception):
                INPUT10=1
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
                
        hvthermalresisderatingfactor = Label(cablesizing,fg='red',text = 'Please Enter Thermal Resistivity Derating Factor= Kr')
        hvthermalresisderatingfactor.place(x=100,y=220)    
        hvinputthermalresis = Text(cablesizing, height = 1.3,width = 5,bg = "light yellow")
        hvinputthermalresis.place(x=420,y=220)    
    
        go10 = Button(cablesizing, height = 1,width = 10,text ="CHECK",command = Take_input10)
        go10.place(x=480,y=217)    
        
    
        def info10display():
            global dbaddr
            global vendor
            img = cv.imread(dbaddr+vendor+'\\'+"Thermal resistivity_HV.jpg",1)
            #print(img)
            #img = cv.resize(img, (950, 690))
            cv.imshow("Thermal Resistivity Derating Factor for HV Cables", img)
            
            cv.waitKey(delay=0)
            cv.destroyAllWindows()
                     
        info10 = Button(cablesizing, height = 1,width = 10,text ="INFO",command = info10display)
        info10.place(x=12,y=217)   
        info10['state']=DISABLED
            
        cb11=Combobox(cablesizing,width=2, values=['A','M'],font=('Times New Roman',12,'bold'))
        cb11.current(0)
        cb11.place(x=570,y=267)    
        def Take_input11():
            global INPUT11
            global cb11a
            try:
                if cb11.get()=='M':
                    INPUT11 = float(hvinputmultigroup.get("1.0", "end-1c"))
                    if 0.95<INPUT11 or INPUT11<0.54:
                        
                        raise Exception
                    #print(type(INPUT11))
                else:
                    INPUT11=float(cb11a.get())
                    
            except(Exception):
                INPUT11=1
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
                
        hvmultigroupderatingfactor = Label(cablesizing,fg='red',text = 'Please Enter MultiCore Group Derating Factor= Kgm')
        hvmultigroupderatingfactor.place(x=100,y=270)    
        hvinputmultigroup = Text(cablesizing, height = 1.3,width = 5,bg = "light yellow")
        hvinputmultigroup.place(x=420,y=270)    
    
        go11 = Button(cablesizing, height = 1,width = 10,text ="CHECK",command = Take_input11)
        go11.place(x=480,y=267)   
        
    
    
        def info11display():
            global dbaddr
            global vendor
            img = cv.imread(dbaddr+vendor+'\\'+"Multicore cables_HV.jpg",1)
            #print(img)
            img = cv.resize(img, (950, 670))
            cv.imshow("Group Derating Factor for HV MultiCore Cables", img)
            
            cv.waitKey(delay=0)
            cv.destroyAllWindows()
                     
        info11 = Button(cablesizing, height = 1,width = 10,text ="INFO",command = info11display)
        info11.place(x=12,y=267)   
        info11['state']=DISABLED
        
        
        
        cb12=Combobox(cablesizing,width=2, values=['A','M'],font=('Times New Roman',12,'bold'))
        cb12.current(0)
        cb12.place(x=570,y=317)
        def Take_input12():
            global INPUT12
            global cb12a
            try:
                if cb12.get()=='M':    
                    INPUT12 = float(hvinputsinglegroup.get("1.0", "end-1c"))
                    if 1<INPUT12 or INPUT12<0.56:
                        
                        raise Exception
                    #print(type(INPUT12))
                else:
                    INPUT12=float(cb12a.get())
            except(Exception):
                INPUT12=1
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
                
        hvsinglegroupderatingfactor = Label(cablesizing,fg='red',text = 'Please Enter SingleCore Group Derating Factor= Kgs')
        hvsinglegroupderatingfactor.place(x=100,y=320)    
        hvinputsinglegroup = Text(cablesizing, height = 1.3,width = 5,bg = "light yellow")
        hvinputsinglegroup.place(x=420,y=320)    
    
        go12 = Button(cablesizing, height = 1,width = 10,text ="CHECK",command = Take_input12)
        go12.place(x=480,y=317)   
        
    
        def info12display():
            global dbaddr
            global vendor
            img = cv.imread(dbaddr+vendor+'\\'+"Single core cables_HV.jpg",1)
            #print(img)
            img = cv.resize(img, (850, 670))
            cv.imshow("Group Derating Factor for HV Single Core Cables", img)
            
            cv.waitKey(delay=0)
            cv.destroyAllWindows()
                 
        info12 = Button(cablesizing, height = 1,width = 10,text ="INFO",command = info12display)
        info12.place(x=12,y=317)  
        info12['state']=DISABLED
    
    
        note = Label(cablesizing, fg='red',bg='yellow',text = 'Note1: Default value for all Derating Factors is 1')
        note.place(x=100,y=367)   
        
        def dest1():
            global calc5
            global calc6
            global calc7
            global calc8
            global form5
            global form6
            global form7
            global form8      
            global reset1
            global browseButton_Excel1
            global calccablesize1
            calc5.destroy()
            calc6.destroy()    
            calc7.destroy()
            calc8.destroy()    
            form5.destroy()    
            form6.destroy() 
            form7.destroy()
            form8.destroy()  
            reset1['state']=DISABLED
            browseButton_Excel1['state']=DISABLED
            if calccablesize1['state']==NORMAL:
                calccablesize1['state']=DISABLED
            go14['state']=NORMAL
            
        reset1 = Button(cablesizing, height = 1,width = 10,text ="RESET",command = dest1)
        reset1.place(x=565,y=367) 
        reset1['state']=DISABLED
        
        def displayDF1():
            global INPUT7
            global INPUT8
            global INPUT9
            global INPUT10
            global INPUT11
            global INPUT12
            global calc5
            global calc6
            global calc7
            global calc8
            global form5
            global form6
            global form7
            global form8   
            global reset1
            global browseButton_Excel1
            #global cb1
            global system
            global Dsa1
            global Dma1
            global Dsg1
            global Dmg1
            global cbalum
            #global go14
            Dsa1=float('{:.2f}'.format(INPUT7*INPUT12))
            Dsg1=float('{:.2f}'.format(INPUT8*INPUT9*INPUT10*INPUT12))
            Dma1=float('{:.2f}'.format(INPUT7*INPUT11))
            Dmg1=float('{:.2f}'.format(INPUT8*INPUT9*INPUT10*INPUT11))
            calc5 = Label(cablesizing, fg='red',text = 'Calculated Value of Derating factor for single Core cables (AG) is: {0}'.format(Dsa1))
            calc5.place(x=100,y=417)  
            form5 = Label(cablesizing, fg='red',text = '(Kta x Kgs)')
            form5.place(x=500,y=417)  
            calc6 = Label(cablesizing, fg='red',text = 'Calculated Value of Derating factor for single Core cables (UG) is: {0}'.format(Dsg1))
            calc6.place(x=100,y=467)   
            form6 = Label(cablesizing, fg='red',text = '(Ktg x Kd x Kr x Kgs)')
            form6.place(x=500,y=467)  
            calc7 = Label(cablesizing, fg='red',text = 'Calculated Value of Derating factor for Multi Core cables (AG) is: {0}'.format(Dma1))
            calc7.place(x=100,y=517)   
            form7 = Label(cablesizing, fg='red',text = '(Kta x Kgm)')
            form7.place(x=500,y=517) 
            calc8 = Label(cablesizing, fg='red',text = 'Calculated Value of Derating factor for Multi Core cables (UG) is: {0}'.format(Dmg1))
            calc8.place(x=100,y=567)  
            form8 = Label(cablesizing, fg='red',text = '(Ktg x Kd x Kr x Kgm)')
            form8.place(x=500,y=567) 
            reset1['state']=NORMAL
            browseButton_Excel1['state']=NORMAL
            go14['state']=DISABLED


        global go14    
        go14 = Button(cablesizing, height = 1,width = 10,text ="CALCULATE",command = displayDF1)
        go14.place(x=480,y=367) 
        go14['state']=DISABLED
            
        def getExcel1 ():
            #global df
            global f1
            global calccablesize1
            global browseButton_Excel1
            global f2
            global cb1
            global cbalum
            global system
            f1=filedialog.askopenfilename()
            calccablesize1['state']=NORMAL
            browseButton_Excel1['state']=DISABLED
            #print(f1[:-9])
            f2=f1.split('/')
            f2=f2[:-1]
            f2='/'.join(f2)
            os.mkdir(f2+'/output')
            #print(f)
            #import_file_path = filedialog.askopenfilename()
            #df = pd.read_excel (import_file_path)
            #print (df)
            if cb1.get()=='EARTHED SYSTEM':
                system='E'
            else:
                system='UE'
            print(system)        
            cbalum=cbalum1.get()    
            print(cbalum)
        #canvas1 = Canvas(cablesizing, width = 300, height = 300, bg = 'lightsteelblue')
        #canvas1.pack()
        #browseButton_Excel = Button(cablesizing,width=16,text='IMPORT EXCEL FILE', command=getExcel, bg='green', fg='white', font=('helvetica', 12, 'bold'))
        #browseButton_Excel.place(x=500, y=620)
        #browseButton_Excel = Button(cablesizing,width=20,fg='blue',text='IMPORT EXCEL FILE', command=getExcel, font=('Times New Roman', 16, 'bold'))
        #browseButton_Excel.place(x=100, y=620)
        browseButton_Excel1 = Button(cablesizing,width=20,fg='blue',text='IMPORT EXCEL FILE', command=getExcel1, font=('Times New Roman', 16, 'bold'))
        browseButton_Excel1.place(x=12, y=620)
        #browseButton_Excel['state']=DISABLED
        browseButton_Excel1['state']=DISABLED
        calccablesize1['state']=DISABLED
        #calccablesize['state']=DISABLED        
        def vendor11select():
            global vendor
            global cb7a
            global cb8a
            global cb9a
            global cb10a
            global cb11a
            global cb12a        
            global go14
            vendor='Polycab'
            vendor11['state']=DISABLED
            vendor12['state']=DISABLED
            vendor13['state']=DISABLED
            resetvendor1['state']=NORMAL 
            info7['state']=NORMAL
            info8['state']=NORMAL
            info9['state']=NORMAL
            info10['state']=NORMAL
            info11['state']=NORMAL
            info12['state']=NORMAL
            cb7a=Combobox(cablesizing,width=4, values=[0.9,0.95,1,1.04,1.1,1.14],font=('Times New Roman',12,'bold'))
            cb7a.current()
            cb7a.place(x=620,y=67)
            cb8a=Combobox(cablesizing,width=4, values=[0.87,0.91,0.96,1,1.03,1.08,1.12],font=('Times New Roman',12,'bold'))
            cb8a.current()
            cb8a.place(x=620,y=117)
            cb9a=Combobox(cablesizing,width=4, values=[0.95,0.96,0.97,0.98,0.99,1],font=('Times New Roman',12,'bold'))
            cb9a.current()
            cb9a.place(x=620,y=167)
            cb10a=Combobox(cablesizing,width=4, values=[0.73,0.8,0.89,1,1.11,1.2],font=('Times New Roman',12,'bold'))
            cb10a.current()
            cb10a.place(x=620,y=217)
            cb11a=Combobox(cablesizing,width=4, values=[0.54,0.58,0.61,0.62,0.65,0.66,0.68,0.69,0.7,0.71,0.72,0.74,0.75,0.76,0.78,0.79,0.8,0.81,0.82,0.83,0.84,0.85,0.86,0.87,0.88,0.89,0.9,0.92,0.93,0.94,0.95,0.96,0.98],font=('Times New Roman',12,'bold'))
            cb11a.current()
            cb11a.place(x=620,y=267)
            cb12a=Combobox(cablesizing,width=4, values=[0.56,0.57,0.6,0.61,0.62,0.65,0.67,0.68,0.71,0.72,0.75,0.76,0.77,0.78,0.79,0.8,0.81,0.82,0.83,0.85,0.86,0.88,0.9,0.92,0.93,0.94,0.96,0.97,0.98,1],font=('Times New Roman',12,'bold'))
            cb12a.current()
            cb12a.place(x=620,y=317)
            go14['state']=NORMAL
            
        def vendor12select():
            global vendor
            global cb7a
            global cb8a
            global cb9a
            global cb10a
            global cb11a
            global cb12a
            global go14
            vendor='KEI'
            vendor11['state']=DISABLED
            vendor12['state']=DISABLED
            vendor13['state']=DISABLED
            resetvendor1['state']=NORMAL
            info7['state']=NORMAL
            info8['state']=NORMAL
            info9['state']=NORMAL
            info10['state']=NORMAL
            info11['state']=NORMAL
            info12['state']=NORMAL
            cb7a=Combobox(cablesizing,width=4, values=[1.16,1.11,1.06,1,0.94,0.88,0.81,0.74],font=('Times New Roman',12,'bold'))
            cb7a.current()
            cb7a.place(x=620,y=67)
            cb8a=Combobox(cablesizing,width=4, values=[1.12,1.08,1.04,1,0.96,0.91,0.87,0.82],font=('Times New Roman',12,'bold'))
            cb8a.current()
            cb8a.place(x=620,y=117)
            cb9a=Combobox(cablesizing,width=4, values=[0.91,0.92,0.93,0.94,0.95,0.97,0.99,1],font=('Times New Roman',12,'bold'))
            cb9a.current()
            cb9a.place(x=620,y=167)
            cb10a=Combobox(cablesizing,width=4, values=['A','M'],font=('Times New Roman',12,'bold'))
            cb10a.current()
            cb10a.place(x=620,y=217)
            cb11a=Combobox(cablesizing,width=4, values=[0.54,0.58,0.61,0.62,0.65,0.66,0.68,0.69,0.7,0.71,0.72,0.73,0.74,0.75,0.76,0.77,0.78,0.79,0.8,0.81,0.82,0.83,0.84,0.85,0.86,0.87,0.88,0.9,0.92,0.93,0.94,0.95,0.96,0.98],font=('Times New Roman',12,'bold'))
            cb11a.current()
            cb11a.place(x=620,y=267)
            cb12a=Combobox(cablesizing,width=4, values=[0.42,0.43,0.44,0.46,0.48,0.5,0.53,0.56,0.57,0.58,0.59,0.6,0.61,0.62,0.63,0.65,0.67,0.68,0.69,0.7,0.71,0.72,0.74,0.75,0.76,0.77,0.78,0.79,0.8,0.81,0.82,0.83,0.84,0.85,0.86,0.87,0.88,0.9,0.92,0.93,0.94,0.96,0.97,0.98,1],font=('Times New Roman',12,'bold'))
            cb12a.current()
            cb12a.place(x=620,y=317)
            go14['state']=NORMAL
            
        def vendor13select():
            global cb7a
            global cb8a
            global cb9a
            global cb10a
            global cb11a
            global cb12a
            global vendor
            global go14
            vendor='Ducab'
            vendor11['state']=DISABLED
            vendor12['state']=DISABLED
            vendor13['state']=DISABLED    
            resetvendor1['state']=NORMAL
            info7['state']=NORMAL
            info8['state']=NORMAL
            info9['state']=NORMAL
            info10['state']=NORMAL
            info11['state']=NORMAL
            info12['state']=NORMAL
            cb7a=Combobox(cablesizing,width=4, values=['A','M'],font=('Times New Roman',12,'bold'))
            cb7a.current()
            cb7a.place(x=620,y=67)
            cb8a=Combobox(cablesizing,width=4, values=['A','M'],font=('Times New Roman',12,'bold'))
            cb8a.current()
            cb8a.place(x=620,y=117)
            cb9a=Combobox(cablesizing,width=4, values=['A','M'],font=('Times New Roman',12,'bold'))
            cb9a.current()
            cb9a.place(x=620,y=167)
            cb10a=Combobox(cablesizing,width=4, values=['A','M'],font=('Times New Roman',12,'bold'))
            cb10a.current()
            cb10a.place(x=620,y=217)
            cb11a=Combobox(cablesizing,width=4, values=['A','M'],font=('Times New Roman',12,'bold'))
            cb11a.current()
            cb11a.place(x=620,y=267)
            cb12a=Combobox(cablesizing,width=4, values=['A','M'],font=('Times New Roman',12,'bold'))
            cb12a.current()
            cb12a.place(x=620,y=317)
            go14['state']=NORMAL            
        
        
        def reset_vendor1():
            global cb7a
            global cb8a
            global cb9a
            global cb10a
            global cb11a
            global cb12a
            global go14
            vendor11['state']=NORMAL
            vendor12['state']=NORMAL
            vendor13['state']=NORMAL
            resetvendor1['state']=DISABLED 
            info7['state']=DISABLED
            info8['state']=DISABLED
            info9['state']=DISABLED
            info10['state']=DISABLED
            info11['state']=DISABLED
            info12['state']=DISABLED
            go14['state']=DISABLED
            cb7a.destroy()
            cb8a.destroy()
            cb9a.destroy()
            cb10a.destroy()
            cb11a.destroy()
            cb12a.destroy()
    
            
        vendor11=Button(cablesizing, width=12,text='POLYCAB',fg='blue',font=('Times New Roman', 16, 'bold'),command=vendor11select)
        vendor11.place(x=750,y=70)
        
        vendor12=Button(cablesizing, width=10,text='KEI',fg='blue',font=('Times New Roman', 16, 'bold'),command=vendor12select)
        vendor12.place(x=937,y=70)
        
        vendor13=Button(cablesizing, width=12,text='DUCAB',fg='blue',font=('Times New Roman', 16, 'bold'),command=vendor13select)
        vendor13.place(x=1100,y=70)
        
        
        
        
        resetvendor1=Button(cablesizing, width=16,text='RESET VENDOR',fg='blue4',font=('Times New Roman', 16, 'bold'),command=lambda:[reset_vendor1(),dest1()])
        resetvendor1.place(x=487,y=620)
        resetvendor1['state']=DISABLED  
        
        systemlabel1=Label(cablesizing,fg='red',text='PLEASE SELECT SYSTEM NEUTRAL GROUNDING')
        systemlabel1.place(x=750, y=140)

        
        
        
        cablesizing.mainloop()   
        
    HVcablesize=Button(mainwindow, width=16,text='HV CABLE SIZING',fg='gray8',font=('Times New Roman', 28, 'bold'),command=LVcable_sizing)
    HVcablesize.place(relx=0.5, rely=0.45, anchor=CENTER)
    
    
    def cable_schedule():
        global cableschedule
        global mainwindow
        mainwindow.iconify()
        cablesize['state']=DISABLED
        cableschedule1['state']=DISABLED
        drumschedule1['state']=DISABLED
        cableschedule=Tk()
        cableschedule.title('CABLE SCHEDULE')
        cableschedule.geometry('1390x680+5+5')
        back=Button(cableschedule, width=16,text='BACK',fg='blue',font=('Times New Roman', 16, 'bold'),command=back2)
        back.place(x=1070,y=620)
        cableschedule.mainloop()
    
    cableschedule1=Button(mainwindow, width=16,text='Cable Schedule',fg='grey8',font=('Times New Roman', 28, 'bold'),command=cable_schedule)
    cableschedule1.place(relx=0.5, rely=0.6, anchor=CENTER)
    
    def drum_schedule():
        global drumschedule
        global mainwindow
        mainwindow.iconify()
        os.system('DrumSchedule.py')
        
        
    drumschedule1=Button(mainwindow, width=16,text='Drum Schedule',fg='grey8',font=('Times New Roman', 28, 'bold'),command=drum_schedule)
    drumschedule1.place(relx=0.5, rely=0.75, anchor=CENTER)
    
    def LVcable_sizing():
        global nrows
        global df
        global f
        global back
        global cablesizing
        global mainwindow
        #global INPUT1
        #global INPUT2
        #global INPUT3
        #global INPUT4
        #global INPUT5
        #global INPUT6
        global INPUT7
        global INPUT8
        global INPUT9
        global INPUT10
        global INPUT11
        global INPUT12
        #global reset
        global reset1
        global browseButton_Excel1
        #global browseButton_Excel
        #global calccablesize
        global calccablesize1
        #global var
        #global var1
        global dbaddr
        dbaddr = askdirectory(title='SELECT DATABASE FOLDER')+'/'
        print(dbaddr)
        def disable_event():
            messagebox.showinfo('ERROR','KINDLY USE EXIT BUTTON')
        
        systeme1=['UN-EARTHED SYSTEM','EARTHED SYSTEM']

        #var = IntVar()
        #var1 = IntVar()
        mainwindow.iconify()
        
        
        HVcablesize['state']=DISABLED
        cableschedule1['state']=DISABLED
        drumschedule1['state']=DISABLED
        LVcablesize['state']=DISABLED
        cablesizing=Tk()
        cablesizing.title('Low Voltage cable sizing')
        cablesizing.geometry('1390x680+5+5')
        cablesizing.configure(bg='lightsteelblue3')
        back=Button(cablesizing, width=10,text='EXIT',fg='blue4',bg='aliceblue',font=('Times New Roman', 16, 'bold'),command=back2)
        back.place(x=732,y=620)
        cablesizing.protocol("WM_DELETE_WINDOW", disable_event)
        global cb1
        cb1=Combobox(cablesizing,width=22, values=systeme1,font=('Times New Roman',12,'bold'))
        cb1.current(0)
        cb1.place(x=1050,y=138)
        #calccablesize=Button(cablesizing, width=10,text='RUN SIZING',fg='blue',font=('Times New Roman', 16, 'bold'),command=calc_cables)
        #calccablesize.place(x=395,y=620)
        calccablesize1=Button(cablesizing, width=10,text='RUN SIZING',bg='aliceblue',fg='blue4',font=('Times New Roman', 16, 'bold'),command=calc_cables1)
        calccablesize1.place(x=307,y=620)
        heading=Label(cablesizing,font=('arial', 30,'bold'),text='LOW VOLTAGE CABLE SIZING',fg='blue4',bg='lightsteelblue3')
        heading.pack()
        #scopeLV=Label(cablesizing,font=('arial', 30,'bold'),text='LV',fg='blue')
        #scopeLV.place(x=100, y=10)
        #scopeHV=Label(cablesizing,font=('arial', 30,'bold'),text='HV',fg='blue')
        #scopeHV.place(x=1180, y=10)
        #R1 = Radiobutton(cablesizing, fg='red', text="Earthed System", variable=var, value=2,command=sel)
        #R1.place(x=180,y=20)
        #R2 = Radiobutton(cablesizing,fg='red', text="Un-Earthed System", variable=var, value=1,command=sel1)
        #R2.place(x=300,y=20)
        #R3 = Radiobutton(cablesizing, fg='red',text="Earthed System", variable=var1, value=2,command=sel2)
        #R3.place(x=910,y=20)
        #R4 = Radiobutton(cablesizing, fg='red',text="Un-Earthed System", variable=var1, value=1,command=sel3)
        #R4.place(x=1030,y=20)
        
        #vendorlabel1=Label(cablesizing,fg='red',text='PLEASE SELECT PREFERRED VENDOR')
        #vendorlabel1.place(x=900, y=70)
        
        def Take_input15():
            global maxmotor
            try:
                maxmotor = float(scmotorrating1.get("1.0", "end-1c"))
                #print(type(INPUT7))
                #print(maxmotor)
            except(Exception):
                
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
        
        scoremotor1=Label(cablesizing,fg='gray8',text='PLEASE PROVIDE MOTOR RATING IN KW \n ABOVE WHICH SINGLE CORE CABLE IS TO BE USED \n DEFAULT RATING IS 2MW')
        scoremotor1.place(x=750, y=190)
        scmotorrating1 = Text(cablesizing, height = 1,width = 5,bg = "light yellow")
        scmotorrating1.place(x=1050,y=190)   
        go15 = Button(cablesizing, height = 1,width = 8,text ="CHECK",command = Take_input15)
        go15.place(x=1115,y=190) 
        #ratingl=Label(cablesizing,fg='red',text='DEFAULT =2MW')
        #ratingl.place(x=1200, y=240)
        
        def del_input15():
            global maxmotor            
            maxmotor=2000   
            scmotorrating1.delete("1.0", "end")
            #print(maxmotor)
        go151 = Button(cablesizing, height = 1,width = 8,text ="RESET",command = del_input15)
        go151.place(x=1190,y=190)
        
        def Take_input16():
            global maxcablesize
            try:
                maxcablesize = float(maxcablesize1.get("1.0", "end-1c"))
                #print(type(INPUT7))
                #print(maxcablesize)
            except(Exception):
                
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
        
        maxcablesize11=Label(cablesizing,fg='gray8',text='PLEASE PROVIDE CABLE SIZE IN SQ.MM. \n ABOVE WHICH SINGLE CORE CABLE IS TO BE USED \n DEFAULT SIZE IS 300 SQ.MM.')
        maxcablesize11.place(x=750, y=260)
        maxcablesize1 = Text(cablesizing, height = 1,width = 5,bg = "light yellow")
        maxcablesize1.place(x=1050,y=260)   
        go16 = Button(cablesizing, height = 1,width = 8,text ="CHECK",command = Take_input16)
        go16.place(x=1115,y=260) 
        #maxsize1=Label(cablesizing,fg='red',text='DEFAULT =300 SQ.MM.')
        #maxsize1.place(x=1250, y=340)
        
        def del_input16():
            global maxcablesize
            maxcablesize=300
            maxcablesize1.delete("1.0", "end")
            #print(maxcablesize)
        
        go161 = Button(cablesizing, height = 1,width = 8,text ="RESET",command = del_input16)
        go161.place(x=1190,y=260) 
        
        alumdecision=Label(cablesizing,fg='gray8',text='DO YOU WANT ALUMINIUM CABLES ONLY?')
        alumdecision.place(x=750, y=330)   
        cbalum1=Combobox(cablesizing,width=4, values=['YES','NO'],font=('Times New Roman',12,'bold'))
        cbalum1.current(0)
        cbalum1.place(x=1050,y=330)
        cbalum=cbalum1.get()
        
        scurrent=Label(cablesizing,fg='gray8',text='PLEASE PROVIDE STARTING CURRENT AS MULTIPLE OF FLC, 20% TOLERANCE WILL BE ADDED INTERNALLY')
        scurrent.place(x=750, y=370)
        
        def Take_input17():
            global factor1
            try:
                factor1 = float(DOL.get("1.0", "end-1c"))
                #print(type(INPUT7))
                #print(maxcablesize)
            except(Exception):
                
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
        
        DOLlabel=Label(cablesizing,fg='gray8',text='DOL \n DEFAULT=5 TIMES')
        DOLlabel.place(x=750, y=410)
        DOL = Text(cablesizing, height = 1,width = 5,bg = "light yellow")
        DOL.place(x=750,y=453)   
        go17 = Button(cablesizing, height = 1,width = 8,text ="CHECK",command = Take_input17)
        go17.place(x=750,y=480) 
        #maxsize1=Label(cablesizing,fg='red',text='DEFAULT =300 SQ.MM.')
        #maxsize1.place(x=1250, y=340)
        
        def del_input17():
            global factor1
            factor1=5
            DOL.delete("1.0", "end")
            #print(maxcablesize)
        
        go171 = Button(cablesizing, height = 1,width = 8,text ="RESET",command = del_input17)
        go171.place(x=750,y=510)
        
        def Take_input18():
            global factor2
            try:
                factor2 = float(VFD.get("1.0", "end-1c"))
                #print(type(INPUT7))
                #print(maxcablesize)
            except(Exception):
                
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
        
        VFDlabel=Label(cablesizing,fg='gray8',text='VFD \n DEFAULT=1.5 TIMES')
        VFDlabel.place(x=900, y=410)
        VFD = Text(cablesizing, height = 1,width = 5,bg = "light yellow")
        VFD.place(x=900,y=453)   
        go18 = Button(cablesizing, height = 1,width = 8,text ="CHECK",command = Take_input18)
        go18.place(x=900,y=480) 
        #maxsize1=Label(cablesizing,fg='red',text='DEFAULT =300 SQ.MM.')
        #maxsize1.place(x=1250, y=340)
        
        def del_input18():
            global factor2
            factor2=1.5
            VFD.delete("1.0", "end")
            #print(maxcablesize)
        
        go181 = Button(cablesizing, height = 1,width = 8,text ="RESET",command = del_input18)
        go181.place(x=900,y=510)
        
        
        def Take_input19():
            global factor3
            try:
                factor3 = float(SS.get("1.0", "end-1c"))
                #print(type(INPUT7))
                #print(maxcablesize)
            except(Exception):
                
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
        
        SSlabel=Label(cablesizing,fg='red',text='SOFT STARTER \n DEFAULT=1 TIMES')
        SSlabel.place(x=1050, y=410)
        SS = Text(cablesizing, height = 1,width = 5,bg = "light yellow")
        SS.place(x=1050,y=453)   
        go19 = Button(cablesizing, height = 1,width = 8,text ="CHECK",command = Take_input19)
        go19.place(x=1050,y=480) 
        #maxsize1=Label(cablesizing,fg='red',text='DEFAULT =300 SQ.MM.')
        #maxsize1.place(x=1250, y=340)
        
        def del_input19():
            global factor3
            factor3=1
            SS.delete("1.0", "end")
            #print(maxcablesize)
        
        go191 = Button(cablesizing, height = 1,width = 8,text ="RESET",command = del_input19)
        go191.place(x=1050,y=510)
        
        
        def Take_input20():
            global factor4
            try:
                factor4 = float(SD.get("1.0", "end-1c"))
                #print(type(INPUT7))
                #print(maxcablesize)
            except(Exception):
                
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
        
        SDlabel=Label(cablesizing,fg='red',text='STAR DELTA \n DEFAULT=3 TIMES')
        SDlabel.place(x=1200, y=410)
        SD = Text(cablesizing, height = 1,width = 5,bg = "light yellow")
        SD.place(x=1200,y=453)   
        go20 = Button(cablesizing, height = 1,width = 8,text ="CHECK",command = Take_input20)
        go20.place(x=1200,y=480) 
        #maxsize1=Label(cablesizing,fg='red',text='DEFAULT =300 SQ.MM.')
        #maxsize1.place(x=1250, y=340)
        
        def del_input20():
            global factor4
            factor4=3
            SD.delete("1.0", "end")
            #print(maxcablesize)
        
        go201 = Button(cablesizing, height = 1,width = 8,text ="RESET",command = del_input20)
        go201.place(x=1200,y=510)

        w = Canvas(cablesizing, width=2, height=500)
        w.create_line(700, 100, 2, 500)
        w.place(x=700,y=80)
        
        
        
        cb7=Combobox(cablesizing,width=2, values=['A','M'],font=('Times New Roman',12,'bold'))
        cb7.current(0)
        cb7.place(x=570,y=67)
        def Take_input7():
            global INPUT7
            global cb7a
            try:
                if cb7.get()=='M':
                    INPUT7 = float(hvinputatemp.get("1.0", "end-1c"))
                    if 1.14<INPUT7 or INPUT7<0.9:
        
                        raise Exception
                    #print(type(INPUT7))
                else:
                    INPUT7=float(cb7a.get())
            except(Exception):
                INPUT7=1
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
                
        Lvatempderatingfactor = Label(cablesizing,fg='red',text = 'Please Enter Temperature Derating Factor (AG)= Kta')
        Lvatempderatingfactor.place(x=100,y=70)    
        Lvinputatemp = Text(cablesizing, height = 1.3,width = 5,bg = "light yellow")
        Lvinputatemp.place(x=420,y=70)    
    
        go7 = Button(cablesizing, height = 1,width = 10,text ="CHECK",command = Take_input7)
        go7.place(x=480,y=67)  
        
        
          
        
        def info7display():
            global dbaddr
            global vendor
            img = cv.imread(dbaddr+vendor+'\\'+"Temperature derating factor_HV air.jpg",1)
            #print(img)
            #img = cv.resize(img, (850, 670))
            cv.imshow("Temperature Derating Factor in Air", img)
            
            cv.waitKey(delay=0)
            cv.destroyAllWindows()
                      
            
        info7 = Button(cablesizing, height = 1,width = 10,text ="INFO",command = info7display)
        info7.place(x=12,y=67)    
        info7['state']=DISABLED
        
        cb8=Combobox(cablesizing,width=2, values=['A','M'],font=('Times New Roman',12,'bold'))
        cb8.current(0)
        cb8.place(x=570,y=117)
        def Take_input8():
            global INPUT8
            #global cb8
            global cb8a
            try:
                if cb8.get()=='M':
                    
                        INPUT8 = float(hvinputgtemp.get("1.0", "end-1c"))
                        #print(type(INPUT8))
                        if 1.12<INPUT8 or INPUT8<0.87:
                            
                            raise Exception
                else:
                    INPUT8=float(cb8a.get())
                    #print('yyay')
            except(Exception):
                INPUT8=1
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
                #print(Exception)
                
        Lvgtempderatingfactor = Label(cablesizing,fg='red',text = 'Please Enter Temperature Derating Factor (UG)= Ktg')
        Lvgtempderatingfactor.place(x=100,y=120)    
        Lvinputgtemp = Text(cablesizing, height = 1.3,width = 5,bg = "light yellow")
        Lvinputgtemp.place(x=420,y=120)    
    
        go8 = Button(cablesizing, height = 1,width = 10,text ="CHECK",command = Take_input8)
        go8.place(x=480,y=117) 
        
        
        def info8display():
            global dbaddr
            global vendor
            img = cv.imread(dbaddr+vendor+'\\'+"Temperature derating factor_HV ground.jpg",1)
            #print(img)
            #img = cv.resize(img, (850, 670))
            cv.imshow("Temperature Derating Factor in Ground", img)
            
            cv.waitKey(delay=0)
            cv.destroyAllWindows()
                 
        info8 = Button(cablesizing, height = 1,width = 10,text ="INFO",command = info8display)
        info8.place(x=12,y=117)  
        info8['state']=DISABLED
        
        cb9=Combobox(cablesizing,width=2, values=['A','M'],font=('Times New Roman',12,'bold'))
        cb9.current(0)
        cb9.place(x=570,y=167)
        def Take_input9():
            global INPUT9
            global cb9a
            try:
                if cb9.get()=='M':
                    INPUT9 = float(hvinputdepth.get("1.0", "end-1c"))
                    #print(type(INPUT9))
                    if 1<INPUT9 or INPUT9<0.96:
                        
                        raise Exception
                else:
                    INPUT9=float(cb9a.get())
            except(Exception):
                INPUT9=1
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
                
        Lvdepthderatingfactor = Label(cablesizing,fg='red',text = 'Please Enter Depth of Laying Derating Factor= Kd')
        Lvdepthderatingfactor.place(x=100,y=170)    
        Lvinputdepth = Text(cablesizing, height = 1.3,width = 5,bg = "light yellow")
        Lvinputdepth.place(x=420,y=170)    
    
        go9 = Button(cablesizing, height = 1,width = 10,text ="CHECK",command = Take_input9)
        go9.place(x=480,y=167)    
        
        
        def info9display():
            global dbaddr
            global vendor
            img = cv.imread(dbaddr+vendor+'\\'+"Laying depth_HV.jpg",1)
            #print(img)
            #img = cv.resize(img, (950, 690))
            cv.imshow("Laying Depth Derating Factor for HV Cables", img)
            
            cv.waitKey(delay=0)
            cv.destroyAllWindows()
                 
        info9 = Button(cablesizing, height = 1,width = 10,text ="INFO",command = info9display)
        info9.place(x=12,y=167) 
        info9['state']=DISABLED
        
        
        cb10=Combobox(cablesizing,width=2, values=['A','M'],font=('Times New Roman',12,'bold'))
        cb10.current(0)
        cb10.place(x=570,y=217)
        def Take_input10():
            global INPUT10
            global cb10a
            try:
                if cb10.get()=='M':
                    INPUT10 = float(hvinputthermalresis.get("1.0", "end-1c"))
                    if 1.2<INPUT10 or INPUT10<0.73:
                        
                        raise Exception
                    #print(type(INPUT10))
                else:
                    INPUT10=float(cb10a.get())
            except(Exception):
                INPUT10=1
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
                
        Lvthermalresisderatingfactor = Label(cablesizing,fg='red',text = 'Please Enter Thermal Resistivity Derating Factor= Kr')
        Lvthermalresisderatingfactor.place(x=100,y=220)    
        Lvinputthermalresis = Text(cablesizing, height = 1.3,width = 5,bg = "light yellow")
        Lvinputthermalresis.place(x=420,y=220)    
    
        go10 = Button(cablesizing, height = 1,width = 10,text ="CHECK",command = Take_input10)
        go10.place(x=480,y=217)    
        
    
        def info10display():
            global dbaddr
            global vendor
            img = cv.imread(dbaddr+vendor+'\\'+"Thermal resistivity_HV.jpg",1)
            #print(img)
            #img = cv.resize(img, (950, 690))
            cv.imshow("Thermal Resistivity Derating Factor for HV Cables", img)
            
            cv.waitKey(delay=0)
            cv.destroyAllWindows()
                     
        info10 = Button(cablesizing, height = 1,width = 10,text ="INFO",command = info10display)
        info10.place(x=12,y=217)   
        info10['state']=DISABLED
            
        cb11=Combobox(cablesizing,width=2, values=['A','M'],font=('Times New Roman',12,'bold'))
        cb11.current(0)
        cb11.place(x=570,y=267)    
        def Take_input11():
            global INPUT11
            global cb11a
            try:
                if cb11.get()=='M':
                    INPUT11 = float(hvinputmultigroup.get("1.0", "end-1c"))
                    if 0.95<INPUT11 or INPUT11<0.54:
                        
                        raise Exception
                    #print(type(INPUT11))
                else:
                    INPUT11=float(cb11a.get())
                    
            except(Exception):
                INPUT11=1
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
                
        Lvmultigroupderatingfactor = Label(cablesizing,fg='red',text = 'Please Enter MultiCore Group Derating Factor= Kgm')
        Lvmultigroupderatingfactor.place(x=100,y=270)    
        Lvinputmultigroup = Text(cablesizing, height = 1.3,width = 5,bg = "light yellow")
        Lvinputmultigroup.place(x=420,y=270)    
    
        go11 = Button(cablesizing, height = 1,width = 10,text ="CHECK",command = Take_input11)
        go11.place(x=480,y=267)   
        
    
    
        def info11display():
            global dbaddr
            global vendor
            img = cv.imread(dbaddr+vendor+'\\'+"Multicore cables_HV.jpg",1)
            #print(img)
            img = cv.resize(img, (950, 670))
            cv.imshow("Group Derating Factor for HV MultiCore Cables", img)
            
            cv.waitKey(delay=0)
            cv.destroyAllWindows()
                     
        info11 = Button(cablesizing, height = 1,width = 10,text ="INFO",command = info11display)
        info11.place(x=12,y=267)   
        info11['state']=DISABLED
        
        
        
        cb12=Combobox(cablesizing,width=2, values=['A','M'],font=('Times New Roman',12,'bold'))
        cb12.current(0)
        cb12.place(x=570,y=317)
        def Take_input12():
            global INPUT12
            global cb12a
            try:
                if cb12.get()=='M':    
                    INPUT12 = float(hvinputsinglegroup.get("1.0", "end-1c"))
                    if 1<INPUT12 or INPUT12<0.56:
                        
                        raise Exception
                    #print(type(INPUT12))
                else:
                    INPUT12=float(cb12a.get())
            except(Exception):
                INPUT12=1
                messagebox.showinfo('ERROR','PLEASE CHECK THE INPUT PROVIDED')
                
        Lvsinglegroupderatingfactor = Label(cablesizing,fg='red',text = 'Please Enter SingleCore Group Derating Factor= Kgs')
        Lvsinglegroupderatingfactor.place(x=100,y=320)    
        Lvinputsinglegroup = Text(cablesizing, height = 1.3,width = 5,bg = "light yellow")
        Lvinputsinglegroup.place(x=420,y=320)    
    
        go12 = Button(cablesizing, height = 1,width = 10,text ="CHECK",command = Take_input12)
        go12.place(x=480,y=317)   
        
    
        def info12display():
            global dbaddr
            global vendor
            img = cv.imread(dbaddr+vendor+'\\'+"Single core cables_HV.jpg",1)
            #print(img)
            img = cv.resize(img, (850, 670))
            cv.imshow("Group Derating Factor for HV Single Core Cables", img)
            
            cv.waitKey(delay=0)
            cv.destroyAllWindows()
                 
        info12 = Button(cablesizing, height = 1,width = 10,text ="INFO",command = info12display)
        info12.place(x=12,y=317)  
        info12['state']=DISABLED
    
    
        note = Label(cablesizing, fg='red',bg='yellow',text = 'Note1: Default value for all Derating Factors is 1')
        note.place(x=100,y=367)   
        
        def dest1():
            global calc5
            global calc6
            global calc7
            global calc8
            global form5
            global form6
            global form7
            global form8      
            global reset1
            global browseButton_Excel1
            global calccablesize1
            calc5.destroy()
            calc6.destroy()    
            calc7.destroy()
            calc8.destroy()    
            form5.destroy()    
            form6.destroy() 
            form7.destroy()
            form8.destroy()  
            reset1['state']=DISABLED
            browseButton_Excel1['state']=DISABLED
            if calccablesize1['state']==NORMAL:
                calccablesize1['state']=DISABLED
            go14['state']=NORMAL
            
        reset1 = Button(cablesizing, height = 1,width = 10,text ="RESET",command = dest1)
        reset1.place(x=565,y=367) 
        reset1['state']=DISABLED
        
        def displayDF1():
            global INPUT7
            global INPUT8
            global INPUT9
            global INPUT10
            global INPUT11
            global INPUT12
            global calc5
            global calc6
            global calc7
            global calc8
            global form5
            global form6
            global form7
            global form8   
            global reset1
            global browseButton_Excel1
            #global cb1
            global system
            global Dsa1
            global Dma1
            global Dsg1
            global Dmg1
            global cbalum
            #global go14
            Dsa1=float('{:.2f}'.format(INPUT7*INPUT12))
            Dsg1=float('{:.2f}'.format(INPUT8*INPUT9*INPUT10*INPUT12))
            Dma1=float('{:.2f}'.format(INPUT7*INPUT11))
            Dmg1=float('{:.2f}'.format(INPUT8*INPUT9*INPUT10*INPUT11))
            calc5 = Label(cablesizing, fg='red',text = 'Calculated Value of Derating factor for single Core cables (AG) is: {0}'.format(Dsa1))
            calc5.place(x=100,y=417)  
            form5 = Label(cablesizing, fg='red',text = '(Kta x Kgs)')
            form5.place(x=500,y=417)  
            calc6 = Label(cablesizing, fg='red',text = 'Calculated Value of Derating factor for single Core cables (UG) is: {0}'.format(Dsg1))
            calc6.place(x=100,y=467)   
            form6 = Label(cablesizing, fg='red',text = '(Ktg x Kd x Kr x Kgs)')
            form6.place(x=500,y=467)  
            calc7 = Label(cablesizing, fg='red',text = 'Calculated Value of Derating factor for Multi Core cables (AG) is: {0}'.format(Dma1))
            calc7.place(x=100,y=517)   
            form7 = Label(cablesizing, fg='red',text = '(Kta x Kgm)')
            form7.place(x=500,y=517) 
            calc8 = Label(cablesizing, fg='red',text = 'Calculated Value of Derating factor for Multi Core cables (UG) is: {0}'.format(Dmg1))
            calc8.place(x=100,y=567)  
            form8 = Label(cablesizing, fg='red',text = '(Ktg x Kd x Kr x Kgm)')
            form8.place(x=500,y=567) 
            reset1['state']=NORMAL
            browseButton_Excel1['state']=NORMAL
            go14['state']=DISABLED


        global go14    
        go14 = Button(cablesizing, height = 1,width = 10,text ="CALCULATE",command = displayDF1)
        go14.place(x=480,y=367) 
        go14['state']=DISABLED
            
        def getExcel1 ():
            #global df
            global f1
            global calccablesize1
            global browseButton_Excel1
            global f2
            global cb1
            global cbalum
            global system
            f1=filedialog.askopenfilename()
            calccablesize1['state']=NORMAL
            browseButton_Excel1['state']=DISABLED
            #print(f1[:-9])
            f2=f1.split('/')
            f2=f2[:-1]
            f2='/'.join(f2)
            os.mkdir(f2+'/output')
            #print(f)
            #import_file_path = filedialog.askopenfilename()
            #df = pd.read_excel (import_file_path)
            #print (df)
            if cb1.get()=='EARTHED SYSTEM':
                system='E'
            else:
                system='UE'
            print(system)        
            cbalum=cbalum1.get()    
            print(cbalum)
        #canvas1 = Canvas(cablesizing, width = 300, height = 300, bg = 'lightsteelblue')
        #canvas1.pack()
        #browseButton_Excel = Button(cablesizing,width=16,text='IMPORT EXCEL FILE', command=getExcel, bg='green', fg='white', font=('helvetica', 12, 'bold'))
        #browseButton_Excel.place(x=500, y=620)
        #browseButton_Excel = Button(cablesizing,width=20,fg='blue',text='IMPORT EXCEL FILE', command=getExcel, font=('Times New Roman', 16, 'bold'))
        #browseButton_Excel.place(x=100, y=620)
        browseButton_Excel1 = Button(cablesizing,width=20,fg='blue4',text='IMPORT EXCEL FILE', command=getExcel1, bg='aliceblue',font=('Times New Roman', 16, 'bold'))
        browseButton_Excel1.place(x=12, y=620)
        #browseButton_Excel['state']=DISABLED
        browseButton_Excel1['state']=DISABLED
        calccablesize1['state']=DISABLED
        #calccablesize['state']=DISABLED        
        def vendor11select():
            global vendor
            global cb7a
            global cb8a
            global cb9a
            global cb10a
            global cb11a
            global cb12a        
            global go14
            vendor='Polycab'
            vendor11['state']=DISABLED
            vendor12['state']=DISABLED
            vendor13['state']=DISABLED
            resetvendor1['state']=NORMAL 
            info7['state']=NORMAL
            info8['state']=NORMAL
            info9['state']=NORMAL
            info10['state']=NORMAL
            info11['state']=NORMAL
            info12['state']=NORMAL
            cb7a=Combobox(cablesizing,width=4, values=[0.9,0.95,1,1.04,1.1,1.14],font=('Times New Roman',12,'bold'))
            cb7a.current()
            cb7a.place(x=620,y=67)
            cb8a=Combobox(cablesizing,width=4, values=[0.87,0.91,0.96,1,1.03,1.08,1.12],font=('Times New Roman',12,'bold'))
            cb8a.current()
            cb8a.place(x=620,y=117)
            cb9a=Combobox(cablesizing,width=4, values=[0.95,0.96,0.97,0.98,0.99,1],font=('Times New Roman',12,'bold'))
            cb9a.current()
            cb9a.place(x=620,y=167)
            cb10a=Combobox(cablesizing,width=4, values=[0.73,0.8,0.89,1,1.11,1.2],font=('Times New Roman',12,'bold'))
            cb10a.current()
            cb10a.place(x=620,y=217)
            cb11a=Combobox(cablesizing,width=4, values=[0.54,0.58,0.61,0.62,0.65,0.66,0.68,0.69,0.7,0.71,0.72,0.74,0.75,0.76,0.78,0.79,0.8,0.81,0.82,0.83,0.84,0.85,0.86,0.87,0.88,0.89,0.9,0.92,0.93,0.94,0.95,0.96,0.98],font=('Times New Roman',12,'bold'))
            cb11a.current()
            cb11a.place(x=620,y=267)
            cb12a=Combobox(cablesizing,width=4, values=[0.56,0.57,0.6,0.61,0.62,0.65,0.67,0.68,0.71,0.72,0.75,0.76,0.77,0.78,0.79,0.8,0.81,0.82,0.83,0.85,0.86,0.88,0.9,0.92,0.93,0.94,0.96,0.97,0.98,1],font=('Times New Roman',12,'bold'))
            cb12a.current()
            cb12a.place(x=620,y=317)
            go14['state']=NORMAL
            
        def vendor12select():
            global vendor
            global cb7a
            global cb8a
            global cb9a
            global cb10a
            global cb11a
            global cb12a
            global go14
            vendor='KEI'
            vendor11['state']=DISABLED
            vendor12['state']=DISABLED
            vendor13['state']=DISABLED
            resetvendor1['state']=NORMAL
            info7['state']=NORMAL
            info8['state']=NORMAL
            info9['state']=NORMAL
            info10['state']=NORMAL
            info11['state']=NORMAL
            info12['state']=NORMAL
            cb7a=Combobox(cablesizing,width=4, values=[1.16,1.11,1.06,1,0.94,0.88,0.81,0.74],font=('Times New Roman',12,'bold'))
            cb7a.current()
            cb7a.place(x=620,y=67)
            cb8a=Combobox(cablesizing,width=4, values=[1.12,1.08,1.04,1,0.96,0.91,0.87,0.82],font=('Times New Roman',12,'bold'))
            cb8a.current()
            cb8a.place(x=620,y=117)
            cb9a=Combobox(cablesizing,width=4, values=[0.91,0.92,0.93,0.94,0.95,0.97,0.99,1],font=('Times New Roman',12,'bold'))
            cb9a.current()
            cb9a.place(x=620,y=167)
            cb10a=Combobox(cablesizing,width=4, values=['A','M'],font=('Times New Roman',12,'bold'))
            cb10a.current()
            cb10a.place(x=620,y=217)
            cb11a=Combobox(cablesizing,width=4, values=[0.54,0.58,0.61,0.62,0.65,0.66,0.68,0.69,0.7,0.71,0.72,0.73,0.74,0.75,0.76,0.77,0.78,0.79,0.8,0.81,0.82,0.83,0.84,0.85,0.86,0.87,0.88,0.9,0.92,0.93,0.94,0.95,0.96,0.98],font=('Times New Roman',12,'bold'))
            cb11a.current()
            cb11a.place(x=620,y=267)
            cb12a=Combobox(cablesizing,width=4, values=[0.42,0.43,0.44,0.46,0.48,0.5,0.53,0.56,0.57,0.58,0.59,0.6,0.61,0.62,0.63,0.65,0.67,0.68,0.69,0.7,0.71,0.72,0.74,0.75,0.76,0.77,0.78,0.79,0.8,0.81,0.82,0.83,0.84,0.85,0.86,0.87,0.88,0.9,0.92,0.93,0.94,0.96,0.97,0.98,1],font=('Times New Roman',12,'bold'))
            cb12a.current()
            cb12a.place(x=620,y=317)
            go14['state']=NORMAL
            
        def vendor13select():
            global cb7a
            global cb8a
            global cb9a
            global cb10a
            global cb11a
            global cb12a
            global vendor
            global go14
            vendor='Ducab'
            vendor11['state']=DISABLED
            vendor12['state']=DISABLED
            vendor13['state']=DISABLED    
            resetvendor1['state']=NORMAL
            info7['state']=NORMAL
            info8['state']=NORMAL
            info9['state']=NORMAL
            info10['state']=NORMAL
            info11['state']=NORMAL
            info12['state']=NORMAL
            cb7a=Combobox(cablesizing,width=4, values=['A','M'],font=('Times New Roman',12,'bold'))
            cb7a.current()
            cb7a.place(x=620,y=67)
            cb8a=Combobox(cablesizing,width=4, values=['A','M'],font=('Times New Roman',12,'bold'))
            cb8a.current()
            cb8a.place(x=620,y=117)
            cb9a=Combobox(cablesizing,width=4, values=['A','M'],font=('Times New Roman',12,'bold'))
            cb9a.current()
            cb9a.place(x=620,y=167)
            cb10a=Combobox(cablesizing,width=4, values=['A','M'],font=('Times New Roman',12,'bold'))
            cb10a.current()
            cb10a.place(x=620,y=217)
            cb11a=Combobox(cablesizing,width=4, values=['A','M'],font=('Times New Roman',12,'bold'))
            cb11a.current()
            cb11a.place(x=620,y=267)
            cb12a=Combobox(cablesizing,width=4, values=['A','M'],font=('Times New Roman',12,'bold'))
            cb12a.current()
            cb12a.place(x=620,y=317)
            go14['state']=NORMAL            
        
        
        def reset_vendor1():
            global cb7a
            global cb8a
            global cb9a
            global cb10a
            global cb11a
            global cb12a
            global go14
            vendor11['state']=NORMAL
            vendor12['state']=NORMAL
            vendor13['state']=NORMAL
            resetvendor1['state']=DISABLED 
            info7['state']=DISABLED
            info8['state']=DISABLED
            info9['state']=DISABLED
            info10['state']=DISABLED
            info11['state']=DISABLED
            info12['state']=DISABLED
            go14['state']=DISABLED
            cb7a.destroy()
            cb8a.destroy()
            cb9a.destroy()
            cb10a.destroy()
            cb11a.destroy()
            cb12a.destroy()
    
            
        vendor11=Button(cablesizing, width=12,text='POLYCAB',fg='blue',font=('Times New Roman', 16, 'bold'),command=vendor11select)
        vendor11.place(x=750,y=70)
        
        vendor12=Button(cablesizing, width=10,text='KEI',fg='blue',font=('Times New Roman', 16, 'bold'),command=vendor12select)
        vendor12.place(x=937,y=70)
        
        vendor13=Button(cablesizing, width=12,text='DUCAB',fg='blue',font=('Times New Roman', 16, 'bold'),command=vendor13select)
        vendor13.place(x=1100,y=70)
        
        
        
        
        resetvendor1=Button(cablesizing, width=16,text='RESET VENDOR',fg='blue4',bg='aliceblue',font=('Times New Roman', 16, 'bold'),command=lambda:[reset_vendor1(),dest1()])
        resetvendor1.place(x=487,y=620)
        resetvendor1['state']=DISABLED  
        
        systemlabel1=Label(cablesizing,fg='red',text='PLEASE SELECT SYSTEM NEUTRAL GROUNDING')
        systemlabel1.place(x=750, y=140)

        
        
        
        cablesizing.mainloop()   
        
    
    
        
    LVcablesize=Button(mainwindow, width=16,text='LV Cable Sizing',fg='gray9',font=('Times New Roman', 28, 'bold'),command=LVcable_sizing)
    LVcablesize.place(relx=0.5, rely=0.3, anchor=CENTER)

    exitb=Button(mainwindow, width=16,text='EXIT',fg='gray8',font=('Times New Roman', 20, 'bold'),command=mainwindow.destroy)
    exitb.place(relx=0.5, rely=0.9, anchor=CENTER)
    
    mainwindow.mainloop()

except Exception as e:
    print (e)